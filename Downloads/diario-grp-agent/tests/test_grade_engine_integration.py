"""Integration tests for the grade engine with realistic GRP-like Excel fixtures.

These fixtures simulate real-world spreadsheet structures that a teacher would
export from the GRP system or build manually for grade entry.
"""
from __future__ import annotations

import pytest
from grp_agent.grade_engine import (
    GradeRule,
    GradeScale,
    GradeSource,
    ScaleConverter,
    block_on_inconsistency,
    combine_grades,
    detect_headers,
    generate_launch_plan,
    generate_preview,
    parse_grades,
    validate_grades,
)
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _create_grp_typical_layout(path, sheet_name="6º ANO ALFA"):
    """Simulate a typical GRP grade export:
    Column A: MATRÍCULA
    Column B: NOME
    Column C: TURMA
    Column D: DISCIPLINA
    Column E: AVALIAÇÃO
    Column F: NOTA 1
    Column G: NOTA 2
    Column H: NOTA FINAL
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {
        1: "MATRÍCULA",
        2: "NOME",
        3: "TURMA",
        4: "DISCIPLINA",
        5: "AVALIAÇÃO",
        6: "NOTA 1",
        7: "NOTA 2",
        8: "NOTA FINAL",
    }
    for col, value in headers.items():
        ws.cell(1, col, value)

    students = [
        (1001, "ANA CAROLINA SILVA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 8.0, 7.5, None),
        (1002, "JOÃO PEDRO SANTOS", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 6.0, 8.0, None),
        (1003, "MARIA EDUARDA LIMA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 9.0, 9.5, None),
        (1004, "LUCAS HENRIQUE ALVES", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 4.0, 5.0, None),
        (1005, "JULIA FERNANDA COSTA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 7.0, 6.5, None),
        (1006, "PEDRO HENRIQUE OLIVEIRA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 10.0, 9.0, None),
        (1007, "ISABELA CRISTINA MARTINS", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 5.5, 7.0, None),
        (1008, "RAFAEL AUGUSTO PEREIRA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 3.0, 4.5, None),
        (1009, "CAMILA SOUZA RODRIGUES", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 8.5, 8.0, None),
        (1010, "THIAGO LUIZ FERREIRA", "6º ANO ALFA", "HISTÓRIA", "2º TRIMESTRE", 6.5, 7.0, None),
    ]

    for row_idx, student in enumerate(students, start=2):
        for col_idx, value in enumerate(student, start=1):
            ws.cell(row_idx, col_idx, value)

    wb.save(path)
    return path


def _create_grp_non_standard_layout(path, sheet_name="7º ANO BETA"):
    """Simulate a spreadsheet with columns in non-standard positions:
    Column A: DISCIPLINA
    Column B: PERÍODO
    Column C: ALUNO(A)
    Column D: ESCOLA
    Column E: CLASSE
    Column F: MATRÍCULA
    Column G: PROVA
    Column H: TRABALHO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {
        1: "DISCIPLINA",
        2: "PERÍODO",
        3: "ALUNO(A)",
        4: "ESCOLA",
        5: "CLASSE",
        6: "MATRÍCULA",
        7: "PROVA",
        8: "TRABALHO",
    }
    for col, value in headers.items():
        ws.cell(1, col, value)

    students = [
        ("MATEMÁTICA", "1º TRIMESTRE", "FERNANDA LOPES", "EMEF CORACI", "7º ANO BETA", 2001, 7.0, 8.0),
        ("MATEMÁTICA", "1º TRIMESTRE", "GABRIEL MENDONÇA", "EMEF CORACI", "7º ANO BETA", 2002, 5.5, 6.0),
        ("MATEMÁTICA", "1º TRIMESTRE", "HELENA CARDOSO", "EMEF CORACI", "7º ANO BETA", 2003, 9.0, 9.5),
        ("MATEMÁTICA", "1º TRIMESTRE", "IGOR VIEIRA", "EMEF CORACI", "7º ANO BETA", 2004, 2.0, 3.0),
        ("MATEMÁTICA", "1º TRIMESTRE", "LAURA PINTO", "EMEF CORACI", "7º ANO BETA", 2005, 8.0, 7.5),
    ]

    for row_idx, student in enumerate(students, start=2):
        for col_idx, value in enumerate(student, start=1):
            ws.cell(row_idx, col_idx, value)

    wb.save(path)
    return path


def _create_grp_with_edge_cases(path, sheet_name="8º ANO"):
    """Simulate a spreadsheet with edge cases:
    - Empty student name
    - Non-numeric grade value
    - Duplicate student entry
    - Grade exceeding maximum
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {
        1: "NOME",
        2: "NOTA",
        3: "MÁXIMA",
    }
    for col, value in headers.items():
        ws.cell(1, col, value)

    rows = [
        ("ALUNO VÁLIDO", 8.0, 10.0),
        ("", 7.0, 10.0),  # Empty name
        ("ALUNO COM TEXTO", "texto", 10.0),  # Non-numeric
        ("ALUNO DUPLICADO", 6.0, 10.0),
        ("ALUNO DUPLICADO", 7.0, 10.0),  # Duplicate
        ("ALUNO ACIMA DO MÁX", 15.0, 10.0),  # Exceeds max
    ]

    for row_idx, student in enumerate(rows, start=2):
        for col_idx, value in enumerate(student, start=1):
            ws.cell(row_idx, col_idx, value)

    wb.save(path)
    return path


def _create_grp_three_source_combination(path, sheet_name="5º ANO"):
    """Simulate a spreadsheet with three grade sources for combination:
    Column A: ESTUDANTE
    Column B: AVALIAÇÃO BIMESTRAL
    Column C: TRABALHO EM GRUPO
    Column D: PARTICIPAÇÃO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {
        1: "ESTUDANTE",
        2: "AVALIAÇÃO BIMESTRAL",
        3: "TRABALHO EM GRUPO",
        4: "PARTICIPAÇÃO",
    }
    for col, value in headers.items():
        ws.cell(1, col, value)

    students = [
        ("CARLOS ALBERTO", 8.0, 7.0, 9.0),
        ("DANIELA FERREIRA", 6.0, 8.0, 7.0),
        ("EDUARDO SANTOS", 9.5, 9.0, 10.0),
        ("FÁTIMA LIMA", 4.0, 5.0, 6.0),
        ("GERALDO SOUZA", 7.5, 6.0, 8.0),
    ]

    for row_idx, student in enumerate(students, start=2):
        for col_idx, value in enumerate(student, start=1):
            ws.cell(row_idx, col_idx, value)

    wb.save(path)
    return path


def _create_grp_scale_conversion(path, sheet_name="4º ANO"):
    """Simulate a spreadsheet with grades in 0-100 scale:
    Column A: NOME
    Column B: NOTA (0-100)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = {1: "NOME", 2: "NOTA"}
    for col, value in headers.items():
        ws.cell(1, col, value)

    students = [
        ("ALUNO NOTA ALTA", 95.0),
        ("ALUNO NOTA MÉDIA", 70.0),
        ("ALUNO NOTA BAIXA", 45.0),
        ("ALUNO NOTA ZERO", 0.0),
        ("ALUNO NOTA MÁXIMA", 100.0),
    ]

    for row_idx, student in enumerate(students, start=2):
        for col_idx, value in enumerate(student, start=1):
            ws.cell(row_idx, col_idx, value)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Integration tests: Header detection
# ---------------------------------------------------------------------------

class TestIntegrationHeaderDetection:
    """Test header detection against realistic GRP-like fixtures."""

    def test_detects_standard_grp_layout(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grp_typical.xlsx")
        result = detect_headers(path)
        assert "aluno" in result.aliases
        assert "turma" in result.aliases
        assert "disciplina" in result.aliases
        assert "periodo" in result.aliases
        assert "matricula" in result.aliases
        assert "fonte_nota" in result.aliases

    def test_detects_non_standard_layout(self, tmp_path):
        path = _create_grp_non_standard_layout(tmp_path / "grp_nonstandard.xlsx")
        result = detect_headers(path)
        assert "aluno" in result.aliases
        assert "turma" in result.aliases
        assert "disciplina" in result.aliases
        assert "periodo" in result.aliases
        assert "escola" in result.aliases
        assert "matricula" in result.aliases
        assert "fonte_nota" in result.aliases

    def test_different_column_positions_detected(self, tmp_path):
        # Standard layout: aluno in col 2
        path1 = _create_grp_typical_layout(tmp_path / "std.xlsx")
        result1 = detect_headers(path1)
        assert result1.aliases["aluno"] == [2]

        # Non-standard layout: aluno in col 3
        path2 = _create_grp_non_standard_layout(tmp_path / "nonstd.xlsx")
        result2 = detect_headers(path2)
        assert result2.aliases["aluno"] == [3]


# ---------------------------------------------------------------------------
# Integration tests: Grade parsing
# ---------------------------------------------------------------------------

class TestIntegrationGradeParsing:
    """Test grade parsing from different layouts."""

    def test_parse_standard_layout(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)  # NOTA 1
        assert len(grades) == 10
        assert grades[0].name == "ANA CAROLINA SILVA"
        assert grades[0].value == 8.0
        assert grades[9].name == "THIAGO LUIZ FERREIRA"
        assert grades[9].value == 6.5

    def test_parse_non_standard_layout(self, tmp_path):
        path = _create_grp_non_standard_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=7)  # PROVA
        assert len(grades) == 5
        assert grades[0].name == "FERNANDA LOPES"
        assert grades[0].value == 7.0
        assert grades[4].name == "LAURA PINTO"
        assert grades[4].value == 8.0

    def test_parse_three_source_layout(self, tmp_path):
        path = _create_grp_three_source_combination(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=2)  # AVALIAÇÃO BIMESTRAL
        assert len(grades) == 5
        assert grades[0].name == "CARLOS ALBERTO"
        assert grades[0].value == 8.0


# ---------------------------------------------------------------------------
# Integration tests: Grade combination
# ---------------------------------------------------------------------------

class TestIntegrationGradeCombination:
    """Test combining grades from multiple sources."""

    def test_combine_two_sources_simple_average(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        nota1 = parse_grades(path, grade_column=6)
        nota2 = parse_grades(path, grade_column=7)

        # Merge by student name
        merged = {}
        for g in nota1:
            merged[g.name] = {"nota1": g}
        for g in nota2:
            if g.name in merged:
                merged[g.name]["nota2"] = g

        rule = GradeRule(name="Média Bimestral", formula="average", sources=["NOTA 1", "NOTA 2"])

        results = []
        for name, sources_dict in merged.items():
            grade_sources = [
                GradeSource(name="NOTA 1", value=sources_dict["nota1"].value, max_value=10.0),
                GradeSource(name="NOTA 2", value=sources_dict["nota2"].value, max_value=10.0),
            ]
            combined = combine_grades(grade_sources, rule)
            results.append((name, combined))

        # Verify combinations
        assert len(results) == 10
        # Ana: (8.0 + 7.5) / 2 = 7.75
        assert results[0][0] == "ANA CAROLINA SILVA"
        assert results[0][1] == pytest.approx(7.75)
        # João: (6.0 + 8.0) / 2 = 7.0
        assert results[1][0] == "JOÃO PEDRO SANTOS"
        assert results[1][1] == pytest.approx(7.0)

    def test_combine_two_sources_weighted_average(self, tmp_path):
        path = _create_grp_non_standard_layout(tmp_path / "grades.xlsx")
        prova = parse_grades(path, grade_column=7)
        trabalho = parse_grades(path, grade_column=8)

        rule = GradeRule(
            name="Média Ponderada",
            formula="weighted_average",
            sources=["PROVA", "TRABALHO"],
        )

        merged = {}
        for g in prova:
            merged[g.name] = {"prova": g}
        for g in trabalho:
            if g.name in merged:
                merged[g.name]["trabalho"] = g

        results = []
        for name, sources_dict in merged.items():
            grade_sources = [
                GradeSource(name="PROVA", value=sources_dict["prova"].value, max_value=10.0, weight=0.6),
                GradeSource(name="TRABALHO", value=sources_dict["trabalho"].value, max_value=10.0, weight=0.4),
            ]
            combined = combine_grades(grade_sources, rule)
            results.append((name, combined))

        # Verify: Fernanda: 7.0*0.6 + 8.0*0.4 = 4.2 + 3.2 = 7.4
        assert results[0][0] == "FERNANDA LOPES"
        assert results[0][1] == pytest.approx(7.4)

    def test_combine_three_sources_average(self, tmp_path):
        path = _create_grp_three_source_combination(tmp_path / "grades.xlsx")
        avaliacao = parse_grades(path, grade_column=2)
        trabalho = parse_grades(path, grade_column=3)
        participacao = parse_grades(path, grade_column=4)

        rule = GradeRule(
            name="Média Geral",
            formula="average",
            sources=["AVALIAÇÃO BIMESTRAL", "TRABALHO EM GRUPO", "PARTICIPAÇÃO"],
        )

        merged = {}
        for g in avaliacao:
            merged[g.name] = {"avaliacao": g}
        for g in trabalho:
            if g.name in merged:
                merged[g.name]["trabalho"] = g
        for g in participacao:
            if g.name in merged:
                merged[g.name]["participacao"] = g

        results = []
        for name, sources_dict in merged.items():
            grade_sources = [
                GradeSource(name="AVALIAÇÃO BIMESTRAL", value=sources_dict["avaliacao"].value, max_value=10.0),
                GradeSource(name="TRABALHO EM GRUPO", value=sources_dict["trabalho"].value, max_value=10.0),
                GradeSource(name="PARTICIPAÇÃO", value=sources_dict["participacao"].value, max_value=10.0),
            ]
            combined = combine_grades(grade_sources, rule)
            results.append((name, combined))

        # Carlos: (8.0 + 7.0 + 9.0) / 3 = 8.0
        assert results[0][0] == "CARLOS ALBERTO"
        assert results[0][1] == pytest.approx(8.0)
        # Eduardo: (9.5 + 9.0 + 10.0) / 3 = 9.5
        assert results[2][0] == "EDUARDO SANTOS"
        assert results[2][1] == pytest.approx(9.5)

    def test_combine_three_sources_weighted(self, tmp_path):
        path = _create_grp_three_source_combination(tmp_path / "grades.xlsx")
        avaliacao = parse_grades(path, grade_column=2)
        trabalho = parse_grades(path, grade_column=3)
        participacao = parse_grades(path, grade_column=4)

        rule = GradeRule(
            name="Média Ponderada Geral",
            formula="weighted_average",
            sources=["AVALIAÇÃO BIMESTRAL", "TRABALHO EM GRUPO", "PARTICIPAÇÃO"],
        )

        merged = {}
        for g in avaliacao:
            merged[g.name] = {"avaliacao": g}
        for g in trabalho:
            if g.name in merged:
                merged[g.name]["trabalho"] = g
        for g in participacao:
            if g.name in merged:
                merged[g.name]["participacao"] = g

        # Verify a specific student with known weights
        sources_dict = merged["CARLOS ALBERTO"]
        grade_sources = [
            GradeSource(name="AVALIAÇÃO BIMESTRAL", value=sources_dict["avaliacao"].value, max_value=10.0, weight=0.5),
            GradeSource(name="TRABALHO EM GRUPO", value=sources_dict["trabalho"].value, max_value=10.0, weight=0.3),
            GradeSource(name="PARTICIPAÇÃO", value=sources_dict["participacao"].value, max_value=10.0, weight=0.2),
        ]
        combined = combine_grades(grade_sources, rule)
        # 8.0*0.5 + 7.0*0.3 + 9.0*0.2 = 4.0 + 2.1 + 1.8 = 7.9
        assert combined == pytest.approx(7.9)


# ---------------------------------------------------------------------------
# Integration tests: Scale conversion
# ---------------------------------------------------------------------------

class TestIntegrationScaleConversion:
    """Test scale conversion with real-world scenarios."""

    def test_convert_0_100_to_0_10(self, tmp_path):
        path = _create_grp_scale_conversion(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=2)

        converter = ScaleConverter(
            from_scale=GradeScale(0, 100),
            to_scale=GradeScale(0, 10),
        )

        results = []
        for g in grades:
            value = float(g.value)
            converted = converter.convert(value)
            results.append((g.name, value, converted))

        # Verify conversions
        assert results[0] == ("ALUNO NOTA ALTA", 95.0, 9.5)
        assert results[1] == ("ALUNO NOTA MÉDIA", 70.0, 7.0)
        assert results[2] == ("ALUNO NOTA BAIXA", 45.0, 4.5)
        assert results[3] == ("ALUNO NOTA ZERO", 0.0, 0.0)
        assert results[4] == ("ALUNO NOTA MÁXIMA", 100.0, 10.0)

    def test_convert_0_10_to_conceptual(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)  # NOTA 1

        converter = ScaleConverter(
            from_scale=GradeScale(0, 10),
            to_scale=GradeScale(0, 10),
            concept_map={0: "I", 5: "R", 7: "B", 8.5: "MB"},
        )

        results = []
        for g in grades:
            value = float(g.value)
            converted = converter.convert(value)
            results.append((g.name, value, converted))

        # Verify conceptual conversions
        # Ana: 8.0 -> "B" (8.0 >= 7, < 8.5)
        assert results[0] == ("ANA CAROLINA SILVA", 8.0, "B")
        # João: 6.0 -> "R" (6.0 >= 5, < 7)
        assert results[1] == ("JOÃO PEDRO SANTOS", 6.0, "R")
        # Maria: 9.0 -> "MB" (9.0 >= 8.5)
        assert results[2] == ("MARIA EDUARDA LIMA", 9.0, "MB")
        # Lucas: 4.0 -> "I" (4.0 < 5)
        assert results[3] == ("LUCAS HENRIQUE ALVES", 4.0, "I")
        # Pedro: 10.0 -> "MB" (10.0 >= 8.5)
        assert results[5] == ("PEDRO HENRIQUE OLIVEIRA", 10.0, "MB")


# ---------------------------------------------------------------------------
# Integration tests: Validation
# ---------------------------------------------------------------------------

class TestIntegrationValidation:
    """Test validation against edge cases in real fixtures."""

    def test_validate_clean_data(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)
        report = validate_grades(grades)
        assert report.blocked == 0
        assert report.valid == 10

    def test_validate_edge_cases(self, tmp_path):
        """Parse skips empty-name rows (by design); validation catches the rest."""
        path = _create_grp_with_edge_cases(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=2)

        # Empty-name row is skipped during parsing, so we get 5 entries
        assert len(grades) == 5

        report = validate_grades(grades)
        assert report.blocked == 3  # non-numeric, duplicate, exceeds-max
        reasons = [item.reason for item in report.items]
        assert "non_numeric_value" in reasons
        assert "duplicate_student" in reasons
        assert "exceeds_maximum" in reasons


# ---------------------------------------------------------------------------
# Integration tests: Missing students
# ---------------------------------------------------------------------------

class TestIntegrationMissingStudents:
    """Test detection of missing students when comparing to GRP."""

    def test_all_students_found(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)
        observed = [g.name.upper() for g in grades]  # All students present
        result = block_on_inconsistency(grades, observed)
        assert result.blocked is False
        assert result.missing_students == []

    def test_some_students_missing(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)
        # Only 5 out of 10 students in GRP
        observed = [g.name.upper() for g in grades[:5]]
        result = block_on_inconsistency(grades, observed)
        assert result.blocked is True
        assert len(result.missing_students) == 5
        assert "missing_students" in result.reasons


# ---------------------------------------------------------------------------
# Integration tests: Launch plan and preview
# ---------------------------------------------------------------------------

class TestIntegrationLaunchPlan:
    """Test launch plan generation with real fixtures."""

    def test_launch_plan_with_scale_conversion(self, tmp_path):
        path = _create_grp_scale_conversion(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=2)
        for g in grades:
            g.max_value = 100.0

        plan = generate_launch_plan(grades, target_scale=GradeScale(0, 10))
        assert plan.blocked is False
        assert len(plan.items) == 5

        # Verify conversions in plan
        assert plan.items[0].student_name == "ALUNO NOTA ALTA"
        assert plan.items[0].target_value == 9.5
        assert plan.items[4].student_name == "ALUNO NOTA MÁXIMA"
        assert plan.items[4].target_value == 10.0

    def test_preview_shows_all_grades(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=6)
        for g in grades:
            g.max_value = 10.0

        preview = generate_preview(grades, target_scale=GradeScale(0, 10))
        assert preview.blocked is False
        assert len(preview.items) == 10
        assert preview.summary["total_students"] == 10

    def test_preview_blocks_on_invalid_data(self, tmp_path):
        path = _create_grp_with_edge_cases(tmp_path / "grades.xlsx")
        grades = parse_grades(path, grade_column=2)
        for g in grades:
            g.max_value = 10.0

        preview = generate_preview(grades, target_scale=GradeScale(0, 10))
        assert preview.blocked is True
        assert len(preview.warnings) > 0


# ---------------------------------------------------------------------------
# Integration tests: End-to-end workflow
# ---------------------------------------------------------------------------

class TestIntegrationEndToEnd:
    """Test the complete workflow: detect → parse → combine → validate → preview."""

    def test_complete_workflow_two_sources(self, tmp_path):
        path = _create_grp_typical_layout(tmp_path / "grades.xlsx")

        # 1. Detect headers
        detection = detect_headers(path)
        assert "aluno" in detection.aliases
        assert "fonte_nota" in detection.aliases

        # 2. Parse grades from two sources
        nota1 = parse_grades(path, grade_column=6)
        nota2 = parse_grades(path, grade_column=7)
        assert len(nota1) == 10
        assert len(nota2) == 10

        # 3. Combine grades
        rule = GradeRule(name="Média Bimestral", formula="average", sources=["NOTA 1", "NOTA 2"])
        merged = {}
        for g in nota1:
            merged[g.name] = {"nota1": g}
        for g in nota2:
            if g.name in merged:
                merged[g.name]["nota2"] = g

        combined_grades = []
        for name, sources_dict in merged.items():
            grade_sources = [
                GradeSource(name="NOTA 1", value=sources_dict["nota1"].value, max_value=10.0),
                GradeSource(name="NOTA 2", value=sources_dict["nota2"].value, max_value=10.0),
            ]
            combined = combine_grades(grade_sources, rule)
            combined_grades.append(GradeSource(name=name, value=combined, max_value=10.0))

        # 4. Validate
        report = validate_grades(combined_grades)
        assert report.blocked == 0
        assert report.valid == 10

        # 5. Preview
        preview = generate_preview(combined_grades, target_scale=GradeScale(0, 10))
        assert preview.blocked is False
        assert len(preview.items) == 10
        assert preview.summary["total_students"] == 10

    def test_complete_workflow_three_sources(self, tmp_path):
        path = _create_grp_three_source_combination(tmp_path / "grades.xlsx")

        # 1. Detect headers
        detection = detect_headers(path)
        assert "aluno" in detection.aliases
        assert "fonte_nota" in detection.aliases

        # 2. Parse grades from three sources
        avaliacao = parse_grades(path, grade_column=2)
        trabalho = parse_grades(path, grade_column=3)
        participacao = parse_grades(path, grade_column=4)

        # 3. Combine grades with weighted average
        rule = GradeRule(
            name="Média Geral",
            formula="weighted_average",
            sources=["AVALIAÇÃO BIMESTRAL", "TRABALHO EM GRUPO", "PARTICIPAÇÃO"],
        )

        merged = {}
        for g in avaliacao:
            merged[g.name] = {"avaliacao": g}
        for g in trabalho:
            if g.name in merged:
                merged[g.name]["trabalho"] = g
        for g in participacao:
            if g.name in merged:
                merged[g.name]["participacao"] = g

        combined_grades = []
        for name, sources_dict in merged.items():
            grade_sources = [
                GradeSource(name="AVALIAÇÃO BIMESTRAL", value=sources_dict["avaliacao"].value, max_value=10.0, weight=0.5),
                GradeSource(name="TRABALHO EM GRUPO", value=sources_dict["trabalho"].value, max_value=10.0, weight=0.3),
                GradeSource(name="PARTICIPAÇÃO", value=sources_dict["participacao"].value, max_value=10.0, weight=0.2),
            ]
            combined = combine_grades(grade_sources, rule)
            combined_grades.append(GradeSource(name=name, value=combined, max_value=10.0))

        # 4. Validate
        report = validate_grades(combined_grades)
        assert report.blocked == 0
        assert report.valid == 5

        # 5. Preview
        preview = generate_preview(combined_grades, target_scale=GradeScale(0, 10))
        assert preview.blocked is False
        assert len(preview.items) == 5

    def test_complete_workflow_with_scale_conversion(self, tmp_path):
        path = _create_grp_scale_conversion(tmp_path / "grades.xlsx")

        # 1. Parse grades (0-100 scale)
        grades = parse_grades(path, grade_column=2)
        for g in grades:
            g.max_value = 100.0

        # 2. Validate
        report = validate_grades(grades)
        assert report.blocked == 0
        assert report.valid == 5

        # 3. Generate launch plan with scale conversion
        plan = generate_launch_plan(grades, target_scale=GradeScale(0, 10))
        assert plan.blocked is False
        assert len(plan.items) == 5

        # 4. Verify conversions
        for item in plan.items:
            assert 0 <= item.target_value <= 10
