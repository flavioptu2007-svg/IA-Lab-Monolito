"""Tests for the GRP audit/preview mode — reads GRP state, compares with Excel,
generates a diff preview, and never writes to GRP.
"""
from __future__ import annotations

from grp_agent.grp_audit import (
    GrpStudentGrade,
    compare_grades,
    generate_audit_preview,
    read_grp_context,
    read_grp_grades,
)
from grp_agent.models import StudentRef
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_grp_page_html(
    school: str = "C. M. E. CORACI MEIRELES DE OLIVEIRA",
    class_name: str = "6º ANO ALFA",
    subject: str = "HISTÓRIA",
    period: str = "2º TRIMESTRE",
    evaluation: str = "AVALIAÇÃO BIMESTRAL",
    students: list[tuple[str, str]] | None = None,
) -> str:
    """Generate minimal GRP HTML with hidden inputs and a student table."""
    if students is None:
        students = [
            ("ANA CAROLINA", "8"),
            ("JOÃO PEDRO", "6"),
            ("MARIA EDUARDA", "9"),
        ]
    rows_html = ""
    for name, grade in students:
        rows_html += (
            f'<tr><td>{name}</td>'
            f'<td><input type="text" value="{grade}"></td></tr>\n'
        )
    return f"""
    <body>
      <input type="hidden" value="{school}">
      <input type="hidden" value="{class_name}">
      <input type="hidden" value="{subject}">
      <input type="hidden" value="{period}">
      <input type="hidden" value="{evaluation}">
      <table><tbody>
        {rows_html}
      </tbody></table>
    </body>
    """


def _make_workbook(
    headers: dict[int, str],
    rows: list[dict[int, object]],
    sheet_name: str = "FOLHA",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, value in headers.items():
        ws.cell(1, col, value)
    for row_idx, row_data in enumerate(rows, start=2):
        for col, value in row_data.items():
            ws.cell(row_idx, col, value)
    return wb


# ---------------------------------------------------------------------------
# 1. Read GRP context from hidden inputs
# ---------------------------------------------------------------------------

class TestReadGrpContext:
    """Extract school, class, subject, period from GRP hidden inputs."""

    def test_extracts_all_fields(self):
        from playwright.sync_api import sync_playwright
        html = _make_grp_page_html()
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            browser.close()
        assert ctx.school == "C. M. E. CORACI MEIRELES DE OLIVEIRA"
        assert ctx.class_name == "6º ANO ALFA"
        assert ctx.subject == "HISTÓRIA"
        assert ctx.period == "2º TRIMESTRE"
        assert ctx.evaluation == "AVALIAÇÃO BIMESTRAL"

    def test_handles_different_values(self):
        from playwright.sync_api import sync_playwright
        html = _make_grp_page_html(
            school="E. M. CACILDA",
            class_name="7º ANO BETA",
            subject="MATEMÁTICA",
            period="1º TRIMESTRE",
            evaluation="PROVA BIMESTRAL",
        )
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            browser.close()
        assert ctx.school == "E. M. CACILDA"
        assert ctx.class_name == "7º ANO BETA"
        assert ctx.subject == "MATEMÁTICA"


# ---------------------------------------------------------------------------
# 2. Read GRP grades from table
# ---------------------------------------------------------------------------

class TestReadGrpGrades:
    """Read student names and grades from the GRP table."""

    def test_reads_all_students(self):
        from playwright.sync_api import sync_playwright
        students = [("ANA", "8"), ("JOÃO", "7"), ("MARIA", "9")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            grades = read_grp_grades(page)
            browser.close()
        assert len(grades) == 3
        assert grades[0].name == "ANA"
        assert grades[0].value == 8.0

    def test_handles_empty_grades(self):
        from playwright.sync_api import sync_playwright
        students = [("ANA", "8"), ("JOÃO", ""), ("MARIA", "9")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            grades = read_grp_grades(page)
            browser.close()
        # Empty grades should still be included (value=None)
        assert len(grades) == 3
        assert grades[1].value is None

    def test_handles_different_students(self):
        from playwright.sync_api import sync_playwright
        students = [("FERNANDA", "7"), ("GABRIEL", "5"), ("HELENA", "10")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            grades = read_grp_grades(page)
            browser.close()
        assert grades[0].name == "FERNANDA"
        assert grades[2].value == 10.0


# ---------------------------------------------------------------------------
# 3. Compare grades (Excel vs GRP)
# ---------------------------------------------------------------------------

class TestCompareGrades:
    """Compare Excel grades with GRP grades and find differences."""

    def test_no_differences(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="JOÃO", registration="1002"),
        ]
        excel_grades = {"ANA": 8.0, "JOÃO": 7.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),
            GrpStudentGrade(name="JOÃO", value=7.0),
        ]
        diff = compare_grades(excel, excel_grades, grp)
        assert diff.items == []
        assert diff.has_changes is False

    def test_detects_new_grade(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="JOÃO", registration="1002"),
        ]
        excel_grades = {"ANA": 8.0, "JOÃO": 7.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),
            GrpStudentGrade(name="JOÃO", value=None),  # Empty in GRP
        ]
        diff = compare_grades(excel, excel_grades, grp)
        assert diff.has_changes is True
        assert len(diff.items) == 1
        assert diff.items[0].student == "JOÃO"
        assert diff.items[0].change_type == "new"

    def test_detects_changed_grade(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
        ]
        excel_grades = {"ANA": 9.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),
        ]
        diff = compare_grades(excel, excel_grades, grp)
        assert diff.has_changes is True
        assert len(diff.items) == 1
        assert diff.items[0].student == "ANA"
        assert diff.items[0].change_type == "changed"
        assert diff.items[0].old_value == 8.0
        assert diff.items[0].new_value == 9.0

    def test_detects_student_not_in_grp(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="FANTASMA", registration="9999"),
        ]
        excel_grades = {"ANA": 8.0, "FANTASMA": 7.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),
        ]
        diff = compare_grades(excel, excel_grades, grp)
        assert diff.has_changes is True
        assert len(diff.items) == 1
        assert diff.items[0].student == "FANTASMA"
        assert diff.items[0].change_type == "not_found"

    def test_detects_student_in_grp_not_in_excel(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
        ]
        excel_grades = {"ANA": 8.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),
            GrpStudentGrade(name="EXTRA", value=5.0),
        ]
        diff = compare_grades(excel, excel_grades, grp)
        # EXTRA is in GRP but not in Excel — should be flagged
        assert diff.has_changes is True
        extra_items = [i for i in diff.items if i.student == "EXTRA"]
        assert len(extra_items) == 1
        assert extra_items[0].change_type == "not_in_source"

    def test_multiple_changes(self):
        excel = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="JOÃO", registration="1002"),
            StudentRef(name="MARIA", registration="1003"),
        ]
        excel_grades = {"ANA": 9.0, "JOÃO": 7.0, "MARIA": 8.0}
        grp = [
            GrpStudentGrade(name="ANA", value=8.0),  # Changed
            GrpStudentGrade(name="JOÃO", value=7.0),  # Same
            GrpStudentGrade(name="MARIA", value=None),  # New
        ]
        diff = compare_grades(excel, excel_grades, grp)
        assert diff.has_changes is True
        assert len(diff.items) == 2  # ANA changed, MARIA new


# ---------------------------------------------------------------------------
# 4. Generate audit preview
# ---------------------------------------------------------------------------

class TestAuditPreview:
    """Generate a human-readable preview of what would change."""

    def test_preview_shows_summary(self):
        from playwright.sync_api import sync_playwright
        students = [("ANA", "8"), ("JOÃO", "6"), ("MARIA", "9")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            grp_grades = read_grp_grades(page)
            browser.close()

        # Simulate Excel data that would change JOÃO's grade
        excel_students = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="JOÃO", registration="1002"),
            StudentRef(name="MARIA", registration="1003"),
        ]
        excel_grades = {"ANA": 8.0, "JOÃO": 8.0, "MARIA": 9.0}

        preview = generate_audit_preview(ctx, excel_students, excel_grades, grp_grades)
        assert preview.context.school == "C. M. E. CORACI MEIRELES DE OLIVEIRA"
        assert preview.total_students == 3
        assert preview.changes_count == 1  # Only JOÃO changed
        assert preview.has_changes is True

    def test_preview_no_changes(self):
        from playwright.sync_api import sync_playwright
        students = [("ANA", "8"), ("JOÃO", "7")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            grp_grades = read_grp_grades(page)
            browser.close()

        excel_students = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="JOÃO", registration="1002"),
        ]
        excel_grades = {"ANA": 8.0, "JOÃO": 7.0}

        preview = generate_audit_preview(ctx, excel_students, excel_grades, grp_grades)
        assert preview.has_changes is False
        assert preview.changes_count == 0

    def test_preview_blocks_when_not_found(self):
        from playwright.sync_api import sync_playwright
        students = [("ANA", "8")]
        html = _make_grp_page_html(students=students)
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            grp_grades = read_grp_grades(page)
            browser.close()

        excel_students = [
            StudentRef(name="ANA", registration="1001"),
            StudentRef(name="FANTASMA", registration="9999"),
        ]
        excel_grades = {"ANA": 8.0, "FANTASMA": 7.0}

        preview = generate_audit_preview(ctx, excel_students, excel_grades, grp_grades)
        assert preview.blocked is True
        assert preview.has_changes is True


# ---------------------------------------------------------------------------
# 5. Integration with grade engine
# ---------------------------------------------------------------------------

class TestIntegrationWithGradeEngine:
    """Test audit mode integration with the grade engine."""

    def test_full_audit_workflow(self, tmp_path):
        from grp_agent.grade_engine import parse_grades
        from playwright.sync_api import sync_playwright

        # Create Excel with grades
        wb = _make_workbook(
            {1: "NOME", 2: "NOTA"},
            [
                {1: "ANA", 2: 8.0},
                {1: "JOÃO", 2: 7.0},
                {1: "MARIA", 2: 9.0},
            ],
        )
        path = tmp_path / "grades.xlsx"
        wb.save(path)

        # Parse Excel
        grades = parse_grades(path, grade_column=2)
        assert len(grades) == 3

        # Simulate GRP page
        grp_students = [("ANA", "8"), ("JOÃO", "6"), ("MARIA", "9")]
        html = _make_grp_page_html(students=grp_students)

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content(html)
            ctx = read_grp_context(page)
            grp_grades = read_grp_grades(page)
            browser.close()

        # Compare
        excel_students = [StudentRef(name=g.name) for g in grades]
        excel_grades_dict = {g.name: float(g.value) for g in grades if isinstance(g.value, (int, float))}

        preview = generate_audit_preview(ctx, excel_students, excel_grades_dict, grp_grades)

        # JOÃO should have changed (6 -> 7)
        assert preview.has_changes is True
        assert preview.changes_count == 1

        # Find the JOÃO change
        joao_change = [i for i in preview.diff.items if i.student == "JOÃO"]
        assert len(joao_change) == 1
        assert joao_change[0].change_type == "changed"
        assert joao_change[0].old_value == 6.0
        assert joao_change[0].new_value == 7.0
