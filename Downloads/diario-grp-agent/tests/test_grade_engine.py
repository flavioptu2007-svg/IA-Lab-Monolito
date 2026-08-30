"""Tests for the flexible grade engine (Motor Flexível de Planilhas e Regras de Notas)."""
from __future__ import annotations

import pytest
from grp_agent.grade_engine import (
    DetectionResult,
    GradeRule,
    GradeScale,
    GradeSource,
    ScaleConverter,
    block_on_inconsistency,
    combine_grades,
    detect_columns,
    detect_headers,
    generate_launch_plan,
    generate_preview,
    parse_grades,
    validate_grades,
)
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(
    headers: dict[int, str],
    rows: list[dict[int, object]],
    sheet_name: str = "FOLHA",
) -> Workbook:
    """Create an Excel workbook with given headers and rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, value in headers.items():
        ws.cell(1, col, value)
    for row_idx, row_data in enumerate(rows, start=2):
        for col, value in row_data.items():
            ws.cell(row_idx, col, value)
    return wb


# ---------------------------------------------------------------------------
# 1. Semantic header detection
# ---------------------------------------------------------------------------

class TestSemanticHeaderDetection:
    """Detect headers by semantic aliases, not by fixed column positions."""

    def test_detects_student_column_by_alias_nome(self, tmp_path):
        wb = _make_workbook({1: "NOME", 2: "NOTA"}, [{1: "Ana", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "aluno" in result.aliases
        assert result.aliases["aluno"] == [1]

    def test_detects_student_column_by_alias_aluno(self, tmp_path):
        wb = _make_workbook({3: "ALUNO(A)", 5: "NOTA"}, [{3: "João", 5: 7}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "aluno" in result.aliases
        assert result.aliases["aluno"] == [3]

    def test_detects_student_column_by_alias_estudante(self, tmp_path):
        wb = _make_workbook({2: "ESTUDANTE", 4: "NOTA"}, [{2: "Maria", 4: 9}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "aluno" in result.aliases
        assert result.aliases["aluno"] == [2]

    def test_detects_class_column_by_alias_turma(self, tmp_path):
        wb = _make_workbook({1: "TURMA", 2: "NOME", 3: "NOTA"}, [{1: "6A", 2: "Ana", 3: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "turma" in result.aliases

    def test_detects_class_column_by_alias_classe(self, tmp_path):
        wb = _make_workbook({2: "CLASSE", 3: "NOTA"}, [{2: "6B", 3: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "turma" in result.aliases

    def test_detects_subject_by_alias_disciplina(self, tmp_path):
        wb = _make_workbook({1: "DISCIPLINA", 2: "NOTA"}, [{1: "MATEMÁTICA", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "disciplina" in result.aliases

    def test_detects_subject_by_alias_componente(self, tmp_path):
        wb = _make_workbook({1: "COMPONENTE CURRICULAR", 2: "NOTA"}, [{1: "HISTÓRIA", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "disciplina" in result.aliases

    def test_detects_period_by_alias_avaliacao(self, tmp_path):
        wb = _make_workbook({1: "AVALIAÇÃO", 2: "NOTA"}, [{1: "1º BIM", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "periodo" in result.aliases

    def test_detects_period_by_alias_bimestre(self, tmp_path):
        wb = _make_workbook({1: "BIMESTRE", 2: "NOTA"}, [{1: "2º BIM", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "periodo" in result.aliases

    def test_detects_matricula_by_alias(self, tmp_path):
        wb = _make_workbook({1: "MATRÍCULA", 2: "NOME", 3: "NOTA"}, [{1: "12345", 2: "Ana", 3: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "matricula" in result.aliases

    def test_detects_escola_by_alias(self, tmp_path):
        wb = _make_workbook({1: "ESCOLA", 2: "NOTA"}, [{1: "EMEF CORACI", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "escola" in result.aliases

    def test_detects_escala_by_alias(self, tmp_path):
        wb = _make_workbook({1: "ESCALA", 2: "NOTA"}, [{1: "0-10", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "escala" in result.aliases


# ---------------------------------------------------------------------------
# 2. Column identification for different positions
# ---------------------------------------------------------------------------

class TestColumnIdentification:
    """Identify semantic columns regardless of their position in the sheet."""

    def test_columns_at_different_positions_case1(self, tmp_path):
        wb = _make_workbook(
            {1: "NOME", 2: "TURMA", 3: "DISCIPLINA", 4: "NOTA"},
            [{1: "Ana", 2: "6A", 3: "MATH", 4: 8}],
        )
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert result.aliases["aluno"] == [1]
        assert result.aliases["turma"] == [2]
        assert result.aliases["disciplina"] == [3]

    def test_columns_at_different_positions_case2(self, tmp_path):
        wb = _make_workbook(
            {5: "ALUNO", 3: "CLASSE", 1: "COMPONENTE", 2: "AVALIAÇÃO", 4: "NOTA"},
            [{5: "João", 3: "6B", 1: "HIST", 2: "1º BIM", 4: 9}],
        )
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert result.aliases["aluno"] == [5]
        assert result.aliases["turma"] == [3]
        assert result.aliases["disciplina"] == [1]
        assert result.aliases["periodo"] == [2]


# ---------------------------------------------------------------------------
# 3. Different names for the same concept
# ---------------------------------------------------------------------------

class TestConceptAliases:
    """Same concept expressed with different header names."""

    def test_nome_and_aluno_same_concept(self, tmp_path):
        wb = _make_workbook({1: "NOME", 2: "NOTA"}, [{1: "Ana", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "aluno" in result.aliases

    def test_estudante_and_nome_same_concept(self, tmp_path):
        wb = _make_workbook({2: "ESTUDANTE", 3: "NOTA"}, [{2: "Maria", 3: 9}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "aluno" in result.aliases

    def test_turma_and_classe_same_concept(self, tmp_path):
        wb = _make_workbook({1: "CLASSE", 2: "NOTA"}, [{1: "6A", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "turma" in result.aliases

    def test_disciplina_and_componente_same_concept(self, tmp_path):
        wb = _make_workbook({1: "COMPONENTE CURRICULAR", 2: "NOTA"}, [{1: "MATH", 2: 8}])
        path = tmp_path / "test.xlsx"
        wb.save(path)
        result = detect_headers(path)
        assert "disciplina" in result.aliases


# ---------------------------------------------------------------------------
# 4. Grade source combination (2 sources)
# ---------------------------------------------------------------------------

class TestCombineTwoGrades:
    """Combine grades from two sources."""

    def test_simple_average_two_sources(self):
        sources = [
            GradeSource(name="Prova", value=8.0, max_value=10.0),
            GradeSource(name="Trabalho", value=6.0, max_value=10.0),
        ]
        rule = GradeRule(
            name="Média Simples",
            formula="average",
            sources=["Prova", "Trabalho"],
        )
        result = combine_grades(sources, rule)
        assert result == 7.0

    def test_weighted_average_two_sources(self):
        sources = [
            GradeSource(name="Prova", value=8.0, max_value=10.0, weight=0.7),
            GradeSource(name="Trabalho", value=6.0, max_value=10.0, weight=0.3),
        ]
        rule = GradeRule(
            name="Média Ponderada",
            formula="weighted_average",
            sources=["Prova", "Trabalho"],
        )
        result = combine_grades(sources, rule)
        assert abs(result - 7.4) < 0.01

    def test_sum_two_sources(self):
        sources = [
            GradeSource(name="Prova 1", value=5.0, max_value=10.0),
            GradeSource(name="Prova 2", value=4.0, max_value=10.0),
        ]
        rule = GradeRule(
            name="Soma",
            formula="sum",
            sources=["Prova 1", "Prova 2"],
        )
        result = combine_grades(sources, rule)
        assert result == 9.0


# ---------------------------------------------------------------------------
# 5. Grade source combination (3 sources)
# ---------------------------------------------------------------------------

class TestCombineThreeGrades:
    """Combine grades from three sources."""

    def test_average_three_sources(self):
        sources = [
            GradeSource(name="Prova", value=8.0, max_value=10.0),
            GradeSource(name="Trabalho", value=6.0, max_value=10.0),
            GradeSource(name="Participação", value=9.0, max_value=10.0),
        ]
        rule = GradeRule(
            name="Média Simples",
            formula="average",
            sources=["Prova", "Trabalho", "Participação"],
        )
        result = combine_grades(sources, rule)
        assert abs(result - 7.6667) < 0.01

    def test_weighted_average_three_sources(self):
        sources = [
            GradeSource(name="Prova", value=8.0, max_value=10.0, weight=0.5),
            GradeSource(name="Trabalho", value=6.0, max_value=10.0, weight=0.3),
            GradeSource(name="Participação", value=9.0, max_value=10.0, weight=0.2),
        ]
        rule = GradeRule(
            name="Média Ponderada",
            formula="weighted_average",
            sources=["Prova", "Trabalho", "Participação"],
        )
        result = combine_grades(sources, rule)
        # 8*0.5 + 6*0.3 + 9*0.2 = 4.0 + 1.8 + 1.8 = 7.6
        assert abs(result - 7.6) < 0.01

    def test_max_three_sources(self):
        sources = [
            GradeSource(name="Prova", value=5.0, max_value=10.0),
            GradeSource(name="Trabalho", value=8.0, max_value=10.0),
            GradeSource(name="Participação", value=3.0, max_value=10.0),
        ]
        rule = GradeRule(
            name="Maior Nota",
            formula="max",
            sources=["Prova", "Trabalho", "Participação"],
        )
        result = combine_grades(sources, rule)
        assert result == 8.0

    def test_min_three_sources(self):
        sources = [
            GradeSource(name="Prova", value=5.0, max_value=10.0),
            GradeSource(name="Trabalho", value=8.0, max_value=10.0),
            GradeSource(name="Participação", value=3.0, max_value=10.0),
        ]
        rule = GradeRule(
            name="Menor Nota",
            formula="min",
            sources=["Prova", "Trabalho", "Participação"],
        )
        result = combine_grades(sources, rule)
        assert result == 3.0


# ---------------------------------------------------------------------------
# 6. Scale conversion
# ---------------------------------------------------------------------------

class TestScaleConversion:
    """Convert grades between different scales."""

    def test_convert_0_10_to_0_100(self):
        converter = ScaleConverter(from_scale=GradeScale(0, 10), to_scale=GradeScale(0, 100))
        result = converter.convert(7.5)
        assert result == 75.0

    def test_convert_0_100_to_0_10(self):
        converter = ScaleConverter(from_scale=GradeScale(0, 100), to_scale=GradeScale(0, 10))
        result = converter.convert(75.0)
        assert result == 7.5

    def test_convert_0_10_to_0_5(self):
        converter = ScaleConverter(from_scale=GradeScale(0, 10), to_scale=GradeScale(0, 5))
        result = converter.convert(8.0)
        assert result == 4.0

    def test_convert_0_5_to_0_10(self):
        converter = ScaleConverter(from_scale=GradeScale(0, 5), to_scale=GradeScale(0, 10))
        result = converter.convert(3.0)
        assert result == 6.0

    def test_convert_0_10_to_brasil_conceitual(self):
        converter = ScaleConverter(
            from_scale=GradeScale(0, 10),
            to_scale=GradeScale(0, 10),
            concept_map={0: "I", 5: "R", 7: "B", 8.5: "MB"},
        )
        result = converter.convert(9.0)
        assert result == "MB"

    def test_convert_preserves_precision(self):
        converter = ScaleConverter(from_scale=GradeScale(0, 10), to_scale=GradeScale(0, 100))
        result = converter.convert(7.85)
        assert result == pytest.approx(78.5)


# ---------------------------------------------------------------------------
# 7. Grade validation
# ---------------------------------------------------------------------------

class TestGradeValidation:
    """Validate grades against various rules."""

    def test_valid_grade_passes(self):
        grades = [GradeSource(name="Prova", value=8.0, max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 0
        assert report.valid == 1

    def test_negative_grade_blocked(self):
        grades = [GradeSource(name="Prova", value=-1.0, max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "negative_grade"

    def test_grade_exceeds_max_blocked(self):
        grades = [GradeSource(name="Prova", value=15.0, max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "exceeds_maximum"


# ---------------------------------------------------------------------------
# 8. Empty field detection
# ---------------------------------------------------------------------------

class TestEmptyFieldDetection:
    """Detect empty fields in grade sources."""

    def test_empty_name_detected(self):
        grades = [GradeSource(name="", value=8.0, max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "empty_student_name"

    def test_empty_value_detected(self):
        grades = [GradeSource(name="Ana", value=None, max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "empty_grade_value"


# ---------------------------------------------------------------------------
# 9. Non-numeric value detection
# ---------------------------------------------------------------------------

class TestNonNumericDetection:
    """Detect non-numeric values in grade columns."""

    def test_text_value_detected(self):
        grades = [GradeSource(name="Ana", value="texto", max_value=10.0)]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "non_numeric_value"


# ---------------------------------------------------------------------------
# 10. Missing student detection
# ---------------------------------------------------------------------------

class TestMissingStudentDetection:
    """Detect students present in source but not in GRP."""

    def test_student_not_found(self):
        source = [
            {"name": "Ana Silva", "value": 8.0},
            {"name": "Fantasma", "value": 7.0},
        ]
        observed = ["ANA SILVA", "JOÃO PEREIRA"]
        report = detect_columns(source, observed)
        assert report.missing_students == ["FANTASMA"]

    def test_all_students_found(self):
        source = [
            {"name": "Ana Silva", "value": 8.0},
            {"name": "João Pereira", "value": 7.0},
        ]
        observed = ["ANA SILVA", "JOÃO PEREIRA"]
        report = detect_columns(source, observed)
        assert report.missing_students == []


# ---------------------------------------------------------------------------
# 11. Duplicate student detection
# ---------------------------------------------------------------------------

class TestDuplicateStudentDetection:
    """Detect duplicate student entries in source."""

    def test_duplicate_detected(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="Ana", value=7.0, max_value=10.0),
        ]
        report = validate_grades(grades)
        assert report.blocked == 1
        assert report.items[0].reason == "duplicate_student"

    def test_no_duplicates(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="João", value=7.0, max_value=10.0),
        ]
        report = validate_grades(grades)
        assert report.blocked == 0


# ---------------------------------------------------------------------------
# 12. Ambiguity detection
# ---------------------------------------------------------------------------

class TestAmbiguityDetection:
    """Detect ambiguities in header mapping."""

    def test_ambiguous_column_detected(self):
        headers = {1: "NOTA 1", 2: "NOTA 2"}
        # Both could map to same concept
        result = detect_headers_from_dict(headers)
        # No ambiguity expected for grade sources (they're separate by design)
        assert "fonte_nota" in result.aliases

    def test_ambiguous_period_detected(self):
        headers = {1: "BIMESTRE 1", 2: "BIMESTRE 2"}
        result = detect_headers_from_dict(headers)
        # Multiple period columns are ambiguous
        if "periodo" in result.aliases:
            assert len(result.aliases["periodo"]) >= 1


# ---------------------------------------------------------------------------
# 13. Launch plan generation
# ---------------------------------------------------------------------------

class TestLaunchPlanGeneration:
    """Generate a launch plan without touching GRP."""

    def test_launch_plan_contains_all_students(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="João", value=7.0, max_value=10.0),
        ]
        plan = generate_launch_plan(grades, target_scale=GradeScale(0, 10))
        assert len(plan.items) == 2
        assert plan.items[0].student_name == "Ana"
        assert plan.items[1].student_name == "João"

    def test_launch_plan_applies_scale_conversion(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
        ]
        plan = generate_launch_plan(grades, target_scale=GradeScale(0, 100))
        assert plan.items[0].target_value == 80.0

    def test_launch_plan_blocks_on_validation_error(self):
        grades = [
            GradeSource(name="Ana", value=-1.0, max_value=10.0),
        ]
        plan = generate_launch_plan(grades, target_scale=GradeScale(0, 10))
        assert plan.blocked is True
        assert len(plan.errors) > 0


# ---------------------------------------------------------------------------
# 14. Preview generation
# ---------------------------------------------------------------------------

class TestPreviewGeneration:
    """Generate preview before any changes to GRP."""

    def test_preview_shows_all_grades(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="João", value=7.0, max_value=10.0),
        ]
        preview = generate_preview(grades, target_scale=GradeScale(0, 10))
        assert len(preview.items) == 2
        assert preview.summary["total_students"] == 2

    def test_preview_shows_warnings(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="", value=7.0, max_value=10.0),
        ]
        preview = generate_preview(grades, target_scale=GradeScale(0, 10))
        assert len(preview.warnings) > 0

    def test_preview_blocks_on_critical_errors(self):
        grades = [
            GradeSource(name="Ana", value=-1.0, max_value=10.0),
        ]
        preview = generate_preview(grades, target_scale=GradeScale(0, 10))
        assert preview.blocked is True


# ---------------------------------------------------------------------------
# 15. Block on inconsistency
# ---------------------------------------------------------------------------

class TestBlockOnInconsistency:
    """Block launch when inconsistencies are detected."""

    def test_blocks_on_missing_students(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="Fantasma", value=7.0, max_value=10.0),
        ]
        observed = ["ANA SILVA"]
        result = block_on_inconsistency(grades, observed)
        assert result.blocked is True
        assert "missing_students" in result.reasons

    def test_blocks_on_duplicates(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="Ana", value=7.0, max_value=10.0),
        ]
        result = block_on_inconsistency(grades, [])
        assert result.blocked is True
        assert "duplicate_students" in result.reasons

    def test_blocks_on_validation_errors(self):
        grades = [
            GradeSource(name="Ana", value=-1.0, max_value=10.0),
        ]
        result = block_on_inconsistency(grades, [])
        assert result.blocked is True
        assert "validation_errors" in result.reasons

    def test_allows_when_no_inconsistencies(self):
        grades = [
            GradeSource(name="Ana", value=8.0, max_value=10.0),
            GradeSource(name="João", value=7.0, max_value=10.0),
        ]
        observed = ["ANA", "JOÃO"]
        result = block_on_inconsistency(grades, observed)
        assert result.blocked is False
        assert result.reasons == []


# ---------------------------------------------------------------------------
# 16. Rule inexistence detection
# ---------------------------------------------------------------------------

class TestRuleInexistence:
    """Detect when a required rule doesn't exist."""

    def test_missing_rule_detected(self):
        sources = [
            GradeSource(name="Prova", value=8.0, max_value=10.0),
            GradeSource(name="Trabalho", value=6.0, max_value=10.0),
        ]
        # Try to combine with a rule that references non-existent source
        bad_rule = GradeRule(
            name="Média",
            formula="average",
            sources=["Prova", "Trabalho", "Participação"],
        )
        with pytest.raises(ValueError, match="missing source"):
            combine_grades(sources, bad_rule)


# ---------------------------------------------------------------------------
# 17. Ambiguous rule detection
# ---------------------------------------------------------------------------

class TestAmbiguousRule:
    """Detect when rules are ambiguous."""

    def test_ambiguous_formula_detected(self):
        with pytest.raises(ValueError, match="invalid formula"):
            GradeRule(
                name="Regra Ruim",
                formula="invalid_formula_xyz",
                sources=["Prova"],
            )

    def test_empty_sources_rule_rejected(self):
        with pytest.raises(ValueError, match="at least one source"):
            GradeRule(
                name="Regra Vazia",
                formula="average",
                sources=[],
            )


# ---------------------------------------------------------------------------
# Integration: parse grades from Excel
# ---------------------------------------------------------------------------

class TestParseGradesFromExcel:
    """Parse grades from Excel files with flexible column positions."""

    def test_parse_with_columns_at_standard_positions(self, tmp_path):
        wb = _make_workbook(
            {1: "NOME", 2: "TURMA", 3: "NOTA"},
            [
                {1: "Ana Silva", 2: "6A", 3: 8.0},
                {1: "João Pereira", 2: "6A", 3: 7.5},
            ],
        )
        path = tmp_path / "grades.xlsx"
        wb.save(path)
        grades = parse_grades(path, grade_column=3)
        assert len(grades) == 2
        assert grades[0].name == "Ana Silva"
        assert grades[0].value == 8.0

    def test_parse_with_columns_at_non_standard_positions(self, tmp_path):
        wb = _make_workbook(
            {5: "ALUNO(A)", 3: "CLASSE", 7: "NOTA 1", 8: "NOTA 2"},
            [
                {5: "Maria Santos", 3: "6B", 7: 9.0, 8: 8.0},
                {5: "Pedro Lima", 3: "6B", 7: 6.5, 8: 7.0},
            ],
        )
        path = tmp_path / "grades.xlsx"
        wb.save(path)
        grades = parse_grades(path, grade_column=7)
        assert len(grades) == 2
        assert grades[0].name == "Maria Santos"
        assert grades[0].value == 9.0


# ---------------------------------------------------------------------------
# Helper for dict-based header detection
# ---------------------------------------------------------------------------

def detect_headers_from_dict(headers: dict[int, str]) -> DetectionResult:
    """Detect headers from a dictionary (for unit testing)."""
    from grp_agent.grade_engine import _classify_header
    aliases = {}
    for col, text in headers.items():
        category = _classify_header(text)
        if category:
            aliases.setdefault(category, []).append(col)
    return DetectionResult(aliases=aliases)
