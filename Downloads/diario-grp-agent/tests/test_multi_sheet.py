"""Tests for multi-sheet Excel support with automatic sheet detection.

The user may have a single Excel file with one sheet per class (e.g.,
"6º ANO ALFA", "6º ANO BETA", "7º ANO"). The engine must:
  - Discover all sheets and their headers
  - Find the correct sheet given context (class, subject, period)
  - Parse grades from the selected sheet
"""
from __future__ import annotations

from grp_agent.grade_engine import (
    detect_headers,
    detect_sheets,
    find_sheet,
    parse_grades,
)
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_multi_sheet_workbook(path):
    """Create a workbook with 3 sheets, each representing a different class."""
    wb = Workbook()

    # Sheet 1: 6º ANO ALFA
    ws1 = wb.active
    ws1.title = "6º ANO ALFA"
    for col, val in {1: "MATRÍCULA", 2: "NOME", 3: "TURMA", 4: "NOTA 1", 5: "NOTA 2"}.items():
        ws1.cell(1, col, val)
    for row, data in enumerate([
        (1001, "ANA CAROLINA", "6º ANO ALFA", 8.0, 7.5),
        (1002, "JOÃO PEDRO", "6º ANO ALFA", 6.0, 8.0),
    ], start=2):
        for col, val in enumerate(data, start=1):
            ws1.cell(row, col, val)

    # Sheet 2: 6º ANO BETA
    ws2 = wb.create_sheet("6º ANO BETA")
    for col, val in {1: "MATRÍCULA", 2: "NOME", 3: "TURMA", 4: "NOTA 1", 5: "NOTA 2"}.items():
        ws2.cell(1, col, val)
    for row, data in enumerate([
        (2001, "MARIA EDUARDA", "6º ANO BETA", 9.0, 9.5),
        (2002, "LUCAS HENRIQUE", "6º ANO BETA", 4.0, 5.0),
    ], start=2):
        for col, val in enumerate(data, start=1):
            ws2.cell(row, col, val)

    # Sheet 3: 7º ANO
    ws3 = wb.create_sheet("7º ANO")
    for col, val in {1: "ALUNO", 2: "PROVA", 3: "TRABALHO"}.items():
        ws3.cell(1, col, val)
    for row, data in enumerate([
        ("FERNANDA", 7.0, 8.0),
        ("GABRIEL", 5.5, 6.0),
    ], start=2):
        for col, val in enumerate(data, start=1):
            ws3.cell(row, col, val)

    wb.save(path)
    return path


def _make_workbook_with_context(path):
    """Create a workbook where sheet names don't match but headers have context."""
    wb = Workbook()

    # Sheet 1: "Folha 1" with HISTÓRIA data
    ws1 = wb.active
    ws1.title = "Folha 1"
    for col, val in {1: "NOME", 2: "DISCIPLINA", 3: "BIMESTRE", 4: "NOTA"}.items():
        ws1.cell(1, col, val)
    for row, data in enumerate([
        ("ANA", "HISTÓRIA", "2º TRIMESTRE", 8.0),
        ("JOÃO", "HISTÓRIA", "2º TRIMESTRE", 6.0),
    ], start=2):
        for col, val in enumerate(data, start=1):
            ws1.cell(row, col, val)

    # Sheet 2: "Folha 2" with MATEMÁTICA data
    ws2 = wb.create_sheet("Folha 2")
    for col, val in {1: "NOME", 2: "DISCIPLINA", 3: "BIMESTRE", 4: "NOTA"}.items():
        ws2.cell(1, col, val)
    for row, data in enumerate([
        ("MARIA", "MATEMÁTICA", "1º TRIMESTRE", 9.0),
        ("PEDRO", "MATEMÁTICA", "1º TRIMESTRE", 7.0),
    ], start=2):
        for col, val in enumerate(data, start=1):
            ws2.cell(row, col, val)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 1. Sheet discovery
# ---------------------------------------------------------------------------

class TestSheetDiscovery:
    """Discover all sheets in a workbook."""

    def test_finds_all_sheets(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        sheets = detect_sheets(path)
        assert len(sheets) == 3
        names = [s.name for s in sheets]
        assert "6º ANO ALFA" in names
        assert "6º ANO BETA" in names
        assert "7º ANO" in names

    def test_each_sheet_has_aliases(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        sheets = detect_sheets(path)
        for sheet in sheets:
            assert "aluno" in sheet.aliases

    def test_sheet_info_has_row_count(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        sheets = detect_sheets(path)
        alfa = next(s for s in sheets if s.name == "6º ANO ALFA")
        assert alfa.row_count == 2  # 2 data rows

    def test_empty_sheet_detected(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "VAZIA"
        ws.cell(1, 1, "NOME")
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        sheets = detect_sheets(path)
        assert len(sheets) == 1
        assert sheets[0].row_count == 0


# ---------------------------------------------------------------------------
# 2. Find sheet by name
# ---------------------------------------------------------------------------

class TestFindSheetByName:
    """Find a specific sheet by exact name."""

    def test_finds_exact_name(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = find_sheet(path, sheet_name="6º ANO BETA")
        assert result is not None
        assert result.name == "6º ANO BETA"

    def test_returns_none_for_unknown_name(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = find_sheet(path, sheet_name="8º ANO GAMMA")
        assert result is None


# ---------------------------------------------------------------------------
# 3. Find sheet by context
# ---------------------------------------------------------------------------

class TestFindByContext:
    """Find the correct sheet based on class/subject/period context."""

    def test_finds_by_class_name(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = find_sheet(path, class_name="7º ANO")
        assert result is not None
        assert result.name == "7º ANO"

    def test_finds_by_partial_class_name(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        # "BETA" should match "6º ANO BETA"
        result = find_sheet(path, class_name="BETA")
        assert result is not None
        assert result.name == "6º ANO BETA"

    def test_finds_by_subject_in_headers(self, tmp_path):
        path = _make_workbook_with_context(tmp_path / "ctx.xlsx")
        result = find_sheet(path, subject="HISTÓRIA")
        assert result is not None
        assert result.name == "Folha 1"

    def test_finds_by_period_in_headers(self, tmp_path):
        path = _make_workbook_with_context(tmp_path / "ctx.xlsx")
        result = find_sheet(path, period="1º TRIMESTRE")
        assert result is not None
        assert result.name == "Folha 2"

    def test_returns_none_when_no_match(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = find_sheet(path, class_name="9º ANO", subject="QUÍMICA")
        assert result is None

    def test_multiple_criteria_narrow_down(self, tmp_path):
        path = _make_workbook_with_context(tmp_path / "ctx.xlsx")
        # Both sheets have NOME and NOTA, but only Folha 2 has MATEMÁTICA
        result = find_sheet(path, subject="MATEMÁTICA", period="1º TRIMESTRE")
        assert result is not None
        assert result.name == "Folha 2"


# ---------------------------------------------------------------------------
# 4. detect_headers with sheet parameter
# ---------------------------------------------------------------------------

class TestDetectHeadersWithSheet:
    """detect_headers should accept an optional sheet name."""

    def test_default_uses_first_sheet(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = detect_headers(path)
        assert "aluno" in result.aliases

    def test_specific_sheet(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        result = detect_headers(path, sheet_name="7º ANO")
        assert "aluno" in result.aliases
        # 7º ANO has "ALUNO" in column 1
        assert result.aliases["aluno"] == [1]

    def test_different_sheet_different_columns(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        r1 = detect_headers(path, sheet_name="6º ANO ALFA")
        r2 = detect_headers(path, sheet_name="7º ANO")
        # Both should detect aluno, but in different positions
        assert "aluno" in r1.aliases
        assert "aluno" in r2.aliases


# ---------------------------------------------------------------------------
# 5. parse_grades with sheet parameter
# ---------------------------------------------------------------------------

class TestParseGradesWithSheet:
    """parse_grades should accept an optional sheet name."""

    def test_parse_specific_sheet(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        grades = parse_grades(path, grade_column=4, sheet_name="6º ANO BETA")
        assert len(grades) == 2
        assert grades[0].name == "MARIA EDUARDA"
        assert grades[0].value == 9.0

    def test_parse_different_sheet_different_data(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        g1 = parse_grades(path, grade_column=4, sheet_name="6º ANO ALFA")
        g2 = parse_grades(path, grade_column=4, sheet_name="6º ANO BETA")
        assert g1[0].name == "ANA CAROLINA"
        assert g2[0].name == "MARIA EDUARDA"

    def test_parse_all_sheets(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")
        sheets = detect_sheets(path)
        all_grades = []
        for sheet in sheets:
            grades = parse_grades(path, grade_column=4, sheet_name=sheet.name)
            all_grades.extend(grades)
        assert len(all_grades) == 4  # 2 from ALFA + 2 from BETA


# ---------------------------------------------------------------------------
# 6. End-to-end: detect → find → parse
# ---------------------------------------------------------------------------

class TestMultiSheetEndToEnd:
    """Complete workflow with multi-sheet workbook."""

    def test_auto_detect_and_parse(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path / "multi.xlsx")

        # 1. Discover sheets
        sheets = detect_sheets(path)
        assert len(sheets) == 3

        # 2. Find the right sheet for "7º ANO"
        target = find_sheet(path, class_name="7º ANO")
        assert target is not None

        # 3. Parse grades from that sheet
        grades = parse_grades(path, grade_column=2, sheet_name=target.name)
        assert len(grades) == 2
        assert grades[0].name == "FERNANDA"
        assert grades[0].value == 7.0

    def test_context_based_selection(self, tmp_path):
        path = _make_workbook_with_context(tmp_path / "ctx.xlsx")

        # Find sheet for MATEMÁTICA
        target = find_sheet(path, subject="MATEMÁTICA")
        assert target is not None

        # Parse from that sheet
        grades = parse_grades(path, grade_column=4, sheet_name=target.name)
        assert len(grades) == 2
        assert grades[0].name == "MARIA"
        assert grades[0].value == 9.0
