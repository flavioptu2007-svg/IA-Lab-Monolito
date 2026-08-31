"""Motor Flexível de Planilhas e Regras de Notas.

Detecta cabeçalhos semânticos, identifica colunas por conceito (não por posição),
combina múltiplas fontes de nota, converte escalas e valida antes de lançar.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Header alias mapping
# ---------------------------------------------------------------------------

_HEADER_ALIASES: dict[str, list[str]] = {
    "aluno": ["aluno", "nome", "estudante"],
    "turma": ["turma", "classe"],
    "escola": ["escola", "unidade"],
    "disciplina": ["disciplina", "componente"],
    "matricula": ["matrícula", "matricula", "registro"],
    "periodo": ["período", "periodo", "bimestre", "avaliação", "avaliacao", "trimestre"],
    "escala": ["escala", "nota máxima", "nota maxima", "nota max", "máxima", "maxima", "max"],
    "fonte_nota": ["nota", "prova", "trabalho", "participação", "participacao"],
    "acertos": ["acertos"],
    "total": ["total"],
    "portugues": ["portug", "português", "portugues"],
    "matematica": ["matem", "matemática", "matematica"],
    "ciencias": ["ciênc", "ciência", "ciencias"],
    "criterios": ["critérios", "criterios"],
}

# Class name patterns: "TURMA: ..." in a cell
_CLASS_PATTERN = re.compile(r"turma[:\s]*(.+)", re.IGNORECASE)
# Period/evaluation patterns: "NOTA AGT DO Xº TRIMESTRE", "PROVA DO CICLO II DO Xº TRIMESTRE"
_PERIOD_PATTERN = re.compile(r"(\d)[ºo]\s*(trimestre|bimestre|semestre)", re.IGNORECASE)
_EVALUATION_PATTERN = re.compile(r"(NOTA\s+AGT|PROVA\s+DO\s+CICLO|AVALIAÇÃO|AVALIACAO)", re.IGNORECASE)

_VALID_FORMULAS = frozenset({"average", "weighted_average", "sum", "max", "min"})

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _normalize(text: str) -> str:
    """Lowercase, collapse whitespace, strip."""
    s = re.sub(r"\s+", " ", text.strip().lower())
    return re.sub(r"\s*\(.*?\)\s*", " ", s).strip()


def _classify_header(text: str) -> str | None:
    """Map a header string to a semantic category, or *None*."""
    normalized = _normalize(text)
    if not normalized:
        return None
    for category, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if normalized == alias or re.match(
                r"^" + re.escape(alias) + r"(\s|$)", normalized
            ):
                return category
    return None


# ---------------------------------------------------------------------------
# Public data classes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SheetInfo:
    """Metadata for a single sheet in a workbook."""

    name: str
    aliases: dict[str, list[int]] = field(default_factory=dict)
    row_count: int = 0


@dataclass(frozen=True)
class DetectionResult:
    """Result of semantic header detection."""

    aliases: dict[str, list[int]] = field(default_factory=dict)


@dataclass(frozen=True)
class GradeScale:
    """Defines a numeric grade range."""

    min_value: float
    max_value: float


@dataclass
class GradeSource:
    """A single grade entry coming from a spreadsheet row."""

    name: str
    value: float | str | None = None
    max_value: float | None = None
    weight: float = 1.0


@dataclass(frozen=True)
class GradeRule:
    """Rule that defines how to combine multiple grade sources."""

    name: str
    formula: str
    sources: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        if self.formula not in _VALID_FORMULAS:
            raise ValueError(f"invalid formula: {self.formula!r}")
        if not self.sources:
            raise ValueError("rule must have at least one source")


@dataclass(frozen=True)
class ScaleConverter:
    """Converts a grade value between two scales, optionally to a concept."""

    from_scale: GradeScale
    to_scale: GradeScale
    concept_map: dict[float, str] | None = None

    def convert(self, value: float) -> float | str:
        if self.concept_map:
            sorted_thresholds = sorted(self.concept_map.keys())
            result = self.concept_map[sorted_thresholds[0]]
            for threshold in sorted_thresholds:
                if value >= threshold:
                    result = self.concept_map[threshold]
            return result
        src_range = self.from_scale.max_value - self.from_scale.min_value
        if src_range == 0:
            return self.to_scale.min_value
        normalized = (value - self.from_scale.min_value) / src_range
        dst_range = self.to_scale.max_value - self.to_scale.min_value
        return self.to_scale.min_value + normalized * dst_range


@dataclass(frozen=True)
class GradeValidationItem:
    """Single validation finding."""

    student_name: str
    reason: str


@dataclass(frozen=True)
class GradeValidationReport:
    """Aggregated validation result."""

    items: list[GradeValidationItem] = field(default_factory=list)
    valid: int = 0
    blocked: int = 0


@dataclass(frozen=True)
class LaunchPlanItem:
    """One row in a launch plan."""

    student_name: str
    source_value: float
    target_value: float | str


@dataclass(frozen=True)
class LaunchPlan:
    """Full plan of what would be written to GRP — *read-only*, never writes."""

    items: list[LaunchPlanItem] = field(default_factory=list)
    blocked: bool = False
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class PreviewItem:
    """One row in a preview report."""

    student_name: str
    original_value: float
    converted_value: float | str


@dataclass(frozen=True)
class PreviewReport:
    """Preview of grades before any GRP interaction."""

    items: list[PreviewItem] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    blocked: bool = False
    summary: dict[str, int] = field(default_factory=dict)


@dataclass(frozen=True)
class InconsistencyResult:
    """Result of inconsistency checks."""

    blocked: bool = False
    reasons: list[str] = field(default_factory=list)
    missing_students: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Core public functions
# ---------------------------------------------------------------------------


def _detect_aliases_from_ws(ws) -> dict[str, list[int]]:
    """Classify headers in the first few rows of a worksheet.

    Scans rows 1-5 to handle real-world spreadsheets where:
    - Row 1 contains section headers (e.g., 'NOTA AGT DO 2º TRIMESTRE')
    - Row 2 contains class names (e.g., 'TURMA: 6º ANO ALFA')
    - Row 3 contains actual column headers (e.g., 'NOME', 'ACERTOS', 'TOTAL')
    """
    aliases: dict[str, list[int]] = {}
    header_row: int | None = None
    # Scan rows 1-5 to find the actual header row
    for row_idx in range(1, min(6, (ws.max_row or 1) + 1)):
        for col in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row_idx, col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            if not text:
                continue
            # Check if this looks like a header row (has known aliases)
            category = _classify_header(text)
            if category:
                if header_row is None:
                    header_row = row_idx
                if row_idx == header_row:
                    aliases.setdefault(category, []).append(col)
    return aliases


def _extract_class_from_ws(ws) -> str:
    """Extract class name from rows 1-2 (e.g., 'TURMA: 6º ANO ALFA')."""
    for row_idx in range(1, min(3, (ws.max_row or 1) + 1)):
        for col in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row_idx, col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            match = _CLASS_PATTERN.search(text)
            if match:
                return match.group(1).strip()
    return ""


def _extract_period_from_ws(ws) -> str:
    """Extract period/evaluation from rows 1-2 (e.g., '2º TRIMESTRE')."""
    for row_idx in range(1, min(3, (ws.max_row or 1) + 1)):
        for col in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row_idx, col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            period_match = _PERIOD_PATTERN.search(text)
            if period_match:
                return period_match.group(0)
    return ""


def _extract_evaluation_from_ws(ws) -> str:
    """Extract evaluation type from rows 1-2 (e.g., 'NOTA AGT')."""
    for row_idx in range(1, min(3, (ws.max_row or 1) + 1)):
        for col in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row_idx, col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            eval_match = _EVALUATION_PATTERN.search(text)
            if eval_match:
                return eval_match.group(0).strip()
    return ""


def _find_header_row(ws) -> int:
    """Find the row that contains the actual column headers.

    Scans rows 1-5 looking for the row with the most known aliases.
    """
    best_row = 1
    best_count = 0
    for row_idx in range(1, min(6, (ws.max_row or 1) + 1)):
        count = 0
        for col in range(1, (ws.max_column or 1) + 1):
            cell_value = ws.cell(row_idx, col).value
            if cell_value is None:
                continue
            text = str(cell_value).strip()
            if not text:
                continue
            category = _classify_header(text)
            if category:
                count += 1
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def _data_row_count(ws) -> int:
    """Count non-empty data rows after the header row."""
    header_row = _find_header_row(ws)
    count = 0
    for row in range(header_row + 1, (ws.max_row or 1) + 1):
        # Check if any cell in this row is non-empty
        has_data = False
        for col in range(1, (ws.max_column or 1) + 1):
            val = ws.cell(row, col).value
            if val is not None and str(val).strip():
                has_data = True
                break
        if has_data:
            count += 1
    return count


def detect_headers(
    path: str | Path, *, sheet_name: str | None = None
) -> DetectionResult:
    """Detect semantic headers in a sheet of an Excel workbook.

    If *sheet_name* is ``None`` the active (first) sheet is used.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return DetectionResult(aliases=_detect_aliases_from_ws(ws))


def detect_sheets(path: str | Path) -> list[SheetInfo]:
    """Return metadata for every sheet in the workbook."""
    wb = load_workbook(path, data_only=True)
    result: list[SheetInfo] = []
    for ws in wb.worksheets:
        aliases = _detect_aliases_from_ws(ws)
        result.append(
            SheetInfo(
                name=ws.title,
                aliases=aliases,
                row_count=_data_row_count(ws),
            )
        )
    return result


def find_sheet(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    class_name: str | None = None,
    subject: str | None = None,
    period: str | None = None,
) -> SheetInfo | None:
    """Find the best matching sheet given optional context criteria.

    Resolution order:
    1. Exact *sheet_name* match.
    2. *class_name* substring match against sheet names.
    3. *subject* / *period* match against header values inside each sheet.
    4. ``None`` if nothing matches.
    """
    sheets = detect_sheets(path)
    if not sheets:
        return None

    # 1. Exact sheet name
    if sheet_name:
        for s in sheets:
            if s.name == sheet_name:
                return s
        return None

    # 2. Class name substring in sheet name
    if class_name:
        norm_cls = _normalize(class_name)
        matches = [
            s for s in sheets if norm_cls in _normalize(s.name)
        ]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            # Return the one with most aliases (best header coverage)
            return max(matches, key=lambda s: sum(len(v) for v in s.aliases.values()))

    # 3. Subject / period match inside header values
    if subject or period:
        wb = load_workbook(path, data_only=True)
        best: SheetInfo | None = None
        best_score = -1
        for s in sheets:
            ws = wb[s.name]
            score = 0
            for row in range(1, min(ws.max_row, 5) + 1):
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row, col).value
                    if val is None:
                        continue
                    norm_val = _normalize(str(val))
                    if subject and _normalize(subject) in norm_val:
                        score += 2
                    if period and _normalize(period) in norm_val:
                        score += 2
            if score > best_score:
                best_score = score
                best = s
        if best_score > 0:
            return best

    # 4. No match
    return None


def detect_columns(
    source: list[dict],
    observed: list[str],
) -> InconsistencyResult:
    """Compare source student names against an observed list (from GRP)."""
    normalized_observed = {name.strip().upper() for name in observed}
    missing: list[str] = []
    for row in source:
        name = str(row["name"]).strip().upper()
        if name not in normalized_observed:
            missing.append(name)
    return InconsistencyResult(
        blocked=len(missing) > 0,
        missing_students=missing,
        reasons=["missing_students"] if missing else [],
    )


def parse_grades(
    path: str | Path,
    grade_column: int,
    *,
    sheet_name: str | None = None,
) -> list[GradeSource]:
    """Read a single grade column from a sheet of an Excel file.

    If *sheet_name* is ``None`` the active (first) sheet is used.
    Optionally reads a max-value column when one is detected via the
    ``escala`` semantic alias.

    Handles multi-row headers: scans rows 1-5 to find the header row,
    then reads data from the row after the header.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = _find_header_row(ws)
    name_col: int | None = None
    max_col: int | None = None
    for col in range(1, (ws.max_column or 1) + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value is None:
            continue
        category = _classify_header(str(cell_value))
        if category == "aluno" and name_col is None:
            name_col = col
        elif category == "escala" and max_col is None:
            max_col = col
    if name_col is None:
        raise ValueError("no student name column found")
    result: list[GradeSource] = []
    for row in range(header_row + 1, (ws.max_row or 1) + 1):
        raw_name = ws.cell(row, name_col).value
        if raw_name is None or str(raw_name).strip() == "":
            continue
        raw_value = ws.cell(row, grade_column).value
        # Skip rows with formula errors like #NAME?
        if raw_value is not None and str(raw_value).strip().startswith("#"):
            continue
        if raw_value is None or str(raw_value).strip() == "":
            continue
        try:
            value: float | str | None = float(raw_value)
        except (ValueError, TypeError):
            value = str(raw_value)
        max_value: float | None = None
        if max_col is not None:
            raw_max = ws.cell(row, max_col).value
            if isinstance(raw_max, (int, float)):
                max_value = float(raw_max)
        result.append(
            GradeSource(
                name=str(raw_name).strip(),
                value=value,
                max_value=max_value,
            )
        )
    return result


def combine_grades(sources: list[GradeSource], rule: GradeRule) -> float:
    """Combine grade sources using the given rule and return a single float."""
    source_names = {s.name for s in sources}
    for required in rule.sources:
        if required not in source_names:
            raise ValueError(f"missing source: {required!r}")

    relevant = [s for s in sources if s.name in rule.sources]

    # All relevant values must be numeric
    values: list[float] = []
    for s in relevant:
        if not isinstance(s.value, (int, float)):
            raise TypeError(f"non-numeric value for source {s.name!r}: {s.value!r}")
        values.append(float(s.value))

    if rule.formula == "average":
        return sum(values) / len(values)
    if rule.formula == "weighted_average":
        total_weight = sum(s.weight for s in relevant)
        if total_weight == 0:
            raise ValueError("total weight cannot be zero")
        return sum(float(s.value) * s.weight for s in relevant) / total_weight
    if rule.formula == "sum":
        return sum(values)
    if rule.formula == "max":
        return max(values)
    if rule.formula == "min":
        return min(values)
    raise ValueError(f"invalid formula: {rule.formula!r}")


def validate_grades(grades: list[GradeSource]) -> GradeValidationReport:
    """Validate a list of grade sources and return a report."""
    items: list[GradeValidationItem] = []
    valid = 0
    blocked = 0

    seen_names: dict[str, bool] = {}
    duplicates: set[str] = set()

    for g in grades:
        if g.name in seen_names:
            duplicates.add(g.name)
        seen_names[g.name] = True

    for g in grades:
        if g.name in duplicates:
            if not any(
                i.student_name == g.name and i.reason == "duplicate_student"
                for i in items
            ):
                items.append(
                    GradeValidationItem(student_name=g.name, reason="duplicate_student")
                )
                blocked += 1
            continue

        if not g.name or g.name.strip() == "":
            items.append(
                GradeValidationItem(
                    student_name=g.name or "(empty)", reason="empty_student_name"
                )
            )
            blocked += 1
        elif g.value is None:
            items.append(
                GradeValidationItem(student_name=g.name, reason="empty_grade_value")
            )
            blocked += 1
        elif isinstance(g.value, str):
            items.append(
                GradeValidationItem(student_name=g.name, reason="non_numeric_value")
            )
            blocked += 1
        elif isinstance(g.value, (int, float)):
            if g.value < 0:
                items.append(
                    GradeValidationItem(student_name=g.name, reason="negative_grade")
                )
                blocked += 1
            elif g.max_value is not None and g.value > g.max_value:
                items.append(
                    GradeValidationItem(
                        student_name=g.name, reason="exceeds_maximum"
                    )
                )
                blocked += 1
            else:
                valid += 1
        else:
            items.append(
                GradeValidationItem(student_name=g.name, reason="non_numeric_value")
            )
            blocked += 1

    return GradeValidationReport(items=items, valid=valid, blocked=blocked)


def generate_launch_plan(
    grades: list[GradeSource],
    target_scale: GradeScale,
) -> LaunchPlan:
    """Build a launch plan with scale conversion — never writes to GRP."""
    report = validate_grades(grades)
    if report.blocked > 0:
        errors = [f"{i.student_name}: {i.reason}" for i in report.items]
        return LaunchPlan(items=[], blocked=True, errors=errors)

    src_max = max(
        (g.max_value for g in grades if g.max_value is not None),
        default=10,
    )
    converter = ScaleConverter(
        from_scale=GradeScale(min_value=0, max_value=src_max),
        to_scale=target_scale,
    )

    items = []
    for g in grades:
        value = float(g.value)  # type: ignore[arg-type]
        items.append(
            LaunchPlanItem(
                student_name=g.name,
                source_value=value,
                target_value=converter.convert(value),
            )
        )
    return LaunchPlan(items=items, blocked=False, errors=[])


def generate_preview(
    grades: list[GradeSource],
    target_scale: GradeScale,
) -> PreviewReport:
    """Generate a human-readable preview — never writes to GRP."""
    report = validate_grades(grades)
    warnings = [f"{i.student_name}: {i.reason}" for i in report.items]

    if report.blocked > 0:
        return PreviewReport(
            items=[],
            warnings=warnings,
            blocked=True,
            summary={"total_students": len(grades)},
        )

    src_max = max(
        (g.max_value for g in grades if g.max_value is not None),
        default=10,
    )
    converter = ScaleConverter(
        from_scale=GradeScale(min_value=0, max_value=src_max),
        to_scale=target_scale,
    )

    items = []
    for g in grades:
        value = float(g.value)  # type: ignore[arg-type]
        items.append(
            PreviewItem(
                student_name=g.name,
                original_value=value,
                converted_value=converter.convert(value),
            )
        )
    return PreviewReport(
        items=items,
        warnings=warnings,
        blocked=False,
        summary={"total_students": len(grades)},
    )


def block_on_inconsistency(
    grades: list[GradeSource],
    observed: list[str],
) -> InconsistencyResult:
    """Return *blocked=True* if any inconsistency is found."""
    reasons: list[str] = []
    all_missing: list[str] = []

    col_result = detect_columns(
        [{"name": g.name, "value": g.value} for g in grades],
        observed,
    )
    if col_result.missing_students:
        reasons.append("missing_students")
        all_missing = col_result.missing_students

    report = validate_grades(grades)
    if any(i.reason == "duplicate_student" for i in report.items):
        reasons.append("duplicate_students")

    if any(i.reason not in ("duplicate_student",) for i in report.items):
        reasons.append("validation_errors")

    return InconsistencyResult(
        blocked=len(reasons) > 0,
        reasons=reasons,
        missing_students=all_missing,
    )
