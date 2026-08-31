#!/usr/bin/env python3
"""Auditoria completa READ-ONLY do GRP — TODAS as escolas da sessão.

Para cada escola da sessão e cada turma (disciplina HISTÓRIA, 2º TRIMESTRE):
lista as avaliações disponíveis e, para cada uma, lê data + grade de notas.
Cruza com a planilha quando a turma está no TURMA_MAP (CORACI).

Status por aluno:
  OK / DIVERGENTE / PREENCHIR_DISPONIVEL / PREENCHIR_SOLICITAR / SENTE_FONTE /
  NAO_ENCONTRADO / MATRICULA_ENCERRADA  (turmas com planilha)
  COM_NOTA / SEM_NOTA / MATRICULA_ENCERRADA           (turmas sem planilha)

NÃO escreve nada. Saída: artifacts/audit/auditoria-grp-<ts>.json + resumo.
"""
from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from grp_agent.browser import BrowserSession
from grp_agent.real_grp_audit import navigate_to_grades_page
from openpyxl import load_workbook

EXCEL_PATH = "/home/flavio/Trabalho/Secretária/PROVAS 2026 2/Notas AGT e Ciclo 2 - Trimestre 2.xlsx"
TARGET_DISCIPLINA = "HISTÓRIA"
TARGET_PERIODO = "2º TRIMESTRE"
GRP_URL = "https://sistemas.paracatu.mg.gov.br/GRP/home/sge/turmas-avaliacoes/resultadoavaliacao"
AVALIACOES = ["CONCEITO", "ATIVIDADE COMPLEMENTAR", "AVALIAÇÃO BIMESTRAL"]
TOL = {"CONCEITO": 0.051, "ATIVIDADE COMPLEMENTAR": 0.051, "AVALIAÇÃO BIMESTRAL": 0.11}
ESCOLA_PRINCIPAL = "CORACI"

TURMA_MAP = {
    "6º ANO A": "6º ANO ALFA",
    "6º ANO B": "6º ANO BETA",
    "6º ANO G": "6º ANO GAMA",
    "7º ANO A": "7º ANO ALFA",
    "7º ANO B": "7º ANO BETA",
    "7º ANO G": "7º ANO GAMA",
    "8º ANO A": "8º ANO ALFA",
}


# ---------------------------------------------------------------------------
# DevExtreme helpers
# ---------------------------------------------------------------------------
def _sel(page, label, val):
    el = page.locator(f"label:has-text('{label}')").first
    if not el.count():
        return
    c = el.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    sb = c.locator("dx-select-box").first
    if not sb.count():
        return
    hi = sb.locator("input[type='hidden']").first
    if hi.count() and val.lower() in (hi.get_attribute("value") or "").lower():
        return
    btn = sb.locator("[role='button']").first
    if btn.count():
        btn.click()
    else:
        vi = sb.locator("input[role='combobox']").first
        if vi.count():
            vi.click()
        else:
            sb.click()
    time.sleep(1)
    page.wait_for_timeout(500)
    opt = page.get_by_text(val, exact=True).filter(visible=True).first
    if opt.count():
        opt.click()
        time.sleep(0.5)


def _dropdown_options(page, label) -> list[str]:
    """Abre o select do label, lê as opções visíveis e fecha (Escape)."""
    el = page.locator(f"label:has-text('{label}')").first
    if not el.count():
        return []
    c = el.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    sb = c.locator("dx-select-box").first
    if not sb.count():
        return []
    btn = sb.locator("[role='button']").first
    if btn.count():
        btn.click()
    else:
        sb.click()
    time.sleep(1.5)
    opts = page.locator(".dx-overlay-wrapper .dx-item:visible, [role='option']:visible")
    texts = []
    for i in range(opts.count()):
        t = opts.nth(i).inner_text().strip()
        if t and t not in texts:
            texts.append(t)
    page.keyboard.press("Escape")
    time.sleep(0.5)
    return texts


def _filtrar(page):
    btn = page.locator("dx-button:has-text('Filtrar')").first
    if btn.count():
        btn.click()
        page.wait_for_load_state("networkidle")
        time.sleep(3)


def read_evaluations_list(page) -> list[str]:
    ev_names = {"conceito", "avaliação bimestral", "atividade complementar"}
    all_hg = page.locator(".dx-datagrid-headers")
    for gi in range(all_hg.count()):
        hg = all_hg.nth(gi)
        pg = hg.locator("xpath=ancestor::dx-data-grid[1]")
        if not pg.count():
            continue
        dr = pg.locator(".dx-datagrid-rowsview .dx-data-row")
        if dr.count() == 0:
            continue
        fc = dr.first.locator("td")
        texts = [fc.nth(i).inner_text().strip() for i in range(fc.count())]
        if any(t.lower() in ev_names for t in texts if t):
            evals = []
            for i in range(dr.count()):
                cells = dr.nth(i).locator("td")
                for ci in range(cells.count()):
                    t = cells.nth(ci).inner_text().strip()
                    if t:
                        evals.append(t)
                        break
            return evals
    return []


def read_student_grid(page) -> list[dict]:
    students: list[dict] = []
    ev_names = {"conceito", "avaliação bimestral", "atividade complementar"}
    all_hg = page.locator(".dx-datagrid-headers")
    for gi in range(all_hg.count()):
        hg = all_hg.nth(gi)
        hc = hg.locator("td[role='columnheader']")
        name_col = reg_col = grade_col = sit_col = None
        for ci in range(hc.count()):
            a = (hc.nth(ci).get_attribute("aria-label") or "").lower()
            if "aluno" in a:
                name_col = ci
            elif "matricula" in a or "matrícula" in a:
                reg_col = ci
            elif "nota" in a:
                grade_col = ci
            elif "situa" in a:
                sit_col = ci
        if name_col is None or grade_col is None:
            continue
        pg = hg.locator("xpath=ancestor::dx-data-grid[1]")
        if not pg.count():
            continue
        dr = pg.locator(".dx-datagrid-rowsview .dx-data-row")
        if dr.count() == 0:
            continue
        fc = dr.first.locator("td")
        if fc.count() > name_col and fc.nth(name_col).inner_text().strip().lower() in ev_names:
            continue
        for i in range(dr.count()):
            row = dr.nth(i)
            cells = row.locator("td")
            if cells.count() <= name_col:
                continue
            name = cells.nth(name_col).inner_text().strip()
            if not name or name.lower() in ("aluno", "nome"):
                continue
            reg = cells.nth(reg_col).inner_text().strip() if reg_col is not None and cells.count() > reg_col else None
            sit = cells.nth(sit_col).inner_text().strip() if sit_col is not None and cells.count() > sit_col else None
            inputs = cells.nth(grade_col).locator("input")
            raw = inputs.first.input_value().strip() if inputs.count() else cells.nth(grade_col).inner_text().strip()
            gv = None
            if raw:
                try:
                    gv = float(raw.replace(",", "."))
                except ValueError:
                    gv = None
            students.append({"name": name, "registration": reg, "grade": gv, "situacao": sit})
        if students:
            break
    return students


def read_date_field(page) -> str | None:
    date_label = page.locator("label:has-text('Data')").first
    if not date_label.count():
        return None
    container = date_label.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    datebox = container.locator("dx-date-box, .dx-datebox").first
    if not datebox.count():
        return None
    date_input = datebox.locator("input[role='combobox'], input.dx-texteditor-input, input[type='text']").first
    if not date_input.count():
        return None
    return date_input.input_value().strip() or None


# ---------------------------------------------------------------------------
# Planilha
# ---------------------------------------------------------------------------
def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", "."))
        except ValueError:
            return None
    return None


def read_excel_sources() -> dict:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    out = {}
    for sheet in TURMA_MAP:
        ws = wb[sheet]
        name_col = None
        for c in range(1, 5):
            v = ws.cell(3, c).value
            if v and "nome" in str(v).lower():
                name_col = c
                break
        if name_col is None:
            name_col = 1 if sheet == "8º ANO A" else 2
        data = {}
        for r in range(4, (ws.max_row or 3) + 1):
            raw = ws.cell(r, name_col).value
            if raw is None or not str(raw).strip():
                continue
            e = _num(ws.cell(r, 5).value)
            g = _num(ws.cell(r, 7).value)
            h = _num(ws.cell(r, 8).value)
            j = _num(ws.cell(r, 10).value)
            eh = round(e + h, 2) if e is not None and h is not None else None
            data[str(raw).strip()] = {"G": g, "J": j, "EH": eh}
        out[sheet] = data
    return out


def _norm(name: str) -> str:
    import re
    import unicodedata
    t = "".join(c for c in unicodedata.normalize("NFKD", name) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", t)).strip().lower()


# ---------------------------------------------------------------------------
# Auditoria
# ---------------------------------------------------------------------------
def nav_to_turma(page, escola: str, turma: str) -> None:
    page.goto(GRP_URL, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle")
    time.sleep(2)
    _sel(page, "Escola", escola)
    time.sleep(1)
    _sel(page, "Turma", turma)
    time.sleep(1)
    _sel(page, "Disciplina", TARGET_DISCIPLINA)
    time.sleep(1)
    _sel(page, "Período de Avaliação", TARGET_PERIODO)
    time.sleep(1)
    _filtrar(page)


def audit_evaluation(page, escola: str, turma: str, ev: str, excel_norm: dict | None, col_key: str | None) -> dict:
    """Navega do zero até a turma, abre a avaliação, lê tudo."""
    nav_to_turma(page, escola, turma)
    ec = page.locator(f".dx-data-row td:has-text('{ev}')").first
    if not ec.count():
        return {"erro": "não aparece na lista filtrada"}
    ec.dblclick()
    time.sleep(3)
    page.wait_for_load_state("networkidle")
    data = read_date_field(page)
    alunos = read_student_grid(page)

    itens = []
    usados = set()
    for a in alunos:
        encerrada = bool(a.get("situacao") and a["situacao"].strip().lower().startswith("encerrad"))
        key = _norm(a["name"])
        fonte = excel_norm.get(key) if excel_norm else None
        esperado = fonte[1][col_key] if fonte else None
        if fonte:
            usados.add(key)
        if encerrada:
            status = "MATRICULA_ENCERRADA"
        elif excel_norm is None:
            status = "COM_NOTA" if a["grade"] is not None else "SEM_NOTA"
        elif a["grade"] is None and esperado is not None:
            status = "PREENCHIR_DISPONIVEL"
        elif a["grade"] is None and esperado is None:
            status = "PREENCHIR_SOLICITAR"
        elif a["grade"] is not None and esperado is None:
            status = "SENTE_FONTE"
        elif abs(a["grade"] - esperado) <= TOL[ev]:
            status = "OK"
        else:
            status = "DIVERGENTE"
        itens.append({
            "aluno_grp": a["name"],
            "matricula": a.get("registration"),
            "situacao": a.get("situacao"),
            "nota_grp": a["grade"],
            "esperado_planilha": esperado,
            "aluno_planilha": fonte[0] if fonte else None,
            "status": status,
        })
    if excel_norm:
        for key, (nome, fontes) in excel_norm.items():
            if key not in usados:
                itens.append({
                    "aluno_grp": None,
                    "aluno_planilha": nome,
                    "esperado_planilha": fontes[col_key],
                    "nota_grp": None,
                    "status": "NAO_ENCONTRADO",
                })
    counts: dict = {}
    for it in itens:
        counts[it["status"]] = counts.get(it["status"], 0) + 1
    return {"data": data, "itens": itens, "resumo": counts}


def audit_turma(page, escola: str, turma: str, sheet: str | None, excel_all: dict) -> dict:
    print(f"\n=== [{escola}] {turma} ===", flush=True)
    result: dict = {"escola": escola, "turma": turma, "avaliacoes": {}}

    nav_to_turma(page, escola, turma)
    disponiveis = read_evaluations_list(page)
    result["avaliacoes_disponiveis"] = disponiveis
    if not disponiveis:
        print("  sem avaliações para HISTÓRIA/2º TRIMESTRE", flush=True)
        result["avaliacoes"] = {}
        return result
    print(f"  avaliações: {disponiveis}", flush=True)

    excel_norm = None
    col_keys = {}
    if sheet is not None:
        excel_norm = {_norm(n): (n, f) for n, f in excel_all[sheet].items()}
        col_keys = {"CONCEITO": "J", "ATIVIDADE COMPLEMENTAR": "G", "AVALIAÇÃO BIMESTRAL": "EH"}

    for ev in AVALIACOES:
        if not any(ev.lower() == d.lower() for d in disponiveis):
            result["avaliacoes"][ev] = {"erro": "não disponível para esta turma"}
            continue
        d = audit_evaluation(page, escola, turma, ev, excel_norm, col_keys.get(ev))
        result["avaliacoes"][ev] = d
        if "erro" in d:
            print(f"  {ev}: {d['erro']}", flush=True)
        else:
            print(f"  {ev}: data={d['data']} | {d['resumo']}", flush=True)
    return result


def main():
    excel_all = read_excel_sources()
    turmas_coraci = {v: k for k, v in TURMA_MAP.items()}  # turma GRP -> sheet
    session = BrowserSession.start(headless=False, state_dir="state/playwright")
    auditoria: dict = {"escolas": {}}
    try:
        page = session.open_grp()
        navigate_to_grades_page(page)
        time.sleep(2)
        if not page.locator("label:has-text('Escola')").count():
            print("🔑 Login necessário — conclua o login na janela (até 3 min)...", flush=True)
            for _ in range(36):
                time.sleep(5)
                try:
                    if page.locator("label:has-text('Escola')").count():
                        break
                    if "resultadoavaliacao" not in page.url:
                        navigate_to_grades_page(page)
                except Exception:  # noqa: BLE001
                    navigate_to_grades_page(page)
            if not page.locator("label:has-text('Escola')").count():
                print("❌ Login não concluído.")
                return

        # 1. Descobre escolas
        escolas = _dropdown_options(page, "Escola")
        print(f"ESCOLAS NA SESSÃO: {escolas}", flush=True)
        if not escolas:
            print("❌ Não consegui listar escolas.")
            return

        for escola in escolas:
            _sel(page, "Escola", escola)
            time.sleep(2)
            turmas = _dropdown_options(page, "Turma")
            print(f"\n##### ESCOLA: {escola} — turmas: {turmas}", flush=True)
            auditoria["escolas"][escola] = {"turmas": {}}
            for turma in turmas:
                sheet = turmas_coraci.get(turma) if escola.upper().find("CORACI") >= 0 else None
                auditoria["escolas"][escola]["turmas"][turma] = audit_turma(
                    page, escola, turma, sheet, excel_all
                )
    finally:
        session.close()

    ts = datetime.now().astimezone().strftime("%Y%m%d-%H%M%S")
    out = Path("artifacts/audit") / f"auditoria-grp-{ts}.json"
    out.write_text(json.dumps(auditoria, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nAuditoria salva em: {out}", flush=True)

    print("\n" + "=" * 100)
    print("RESUMO CONSOLIDADO")
    print("=" * 100)
    for escola, dados in auditoria["escolas"].items():
        print(f"\n########## {escola}")
        for turma, td in dados["turmas"].items():
            print(f"\n### {turma} — disponíveis: {td.get('avaliacoes_disponiveis', [])}")
            for ev, d in td["avaliacoes"].items():
                if "erro" in d:
                    print(f"  {ev}: {d['erro']}")
                    continue
                print(f"  {ev} (data {d['data'] or 'SEM'}): " +
                      " · ".join(f"{k}:{v}" for k, v in sorted(d["resumo"].items())))
                for i in d["itens"]:
                    if i["status"] in ("OK", "COM_NOTA"):
                        continue
                    nome = i.get("aluno_grp") or i.get("aluno_planilha")
                    print(f"     [{i['status']}] {nome} — GRP={i.get('nota_grp')} planilha={i.get('esperado_planilha')}")


if __name__ == "__main__":
    main()
