#!/usr/bin/env python3
"""Auditoria OBJETIVA read-only do GRP.

Estratégia (conforme orientação do usuário):
- O formato se repete entre turmas da mesma escola => amostra 1 turma por escola.
- CORACI: completa só o que falta nas notas (BIMESTRAL de 7º BETA/GAMA/8º ALFA).
- CACILDA / CAIC: amostra 1 turma (lista de avaliações + estado geral).
- Depois muda de campo: explora a área de AULAS (lançamento de aula, dias variam).

NÃO escreve nada.
"""
from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from grp_agent.browser import BrowserSession
from grp_agent.real_grp_audit import navigate_to_grades_page
from openpyxl import load_workbook

EXCEL_PATH = "/home/flavio/Trabalho/Secretária/PROVAS 2026 2/Notas AGT e Ciclo 2 - Trimestre 2.xlsx"
TARGET_DISCIPLINA = "HISTÓRIA"
TARGET_PERIODO = "2º TRIMESTRE"
GRP_URL = "https://sistemas.paracatu.mg.gov.br/GRP/home/sge/turmas-avaliacoes/resultadoavaliacao"
GRP_HOME = "https://sistemas.paracatu.mg.gov.br/GRP/"
TOL_BIM = 0.11

ESCOLA_CORACI = "CORACI"
BIMESTRAL_RESTO = ["7º ANO BETA", "7º ANO GAMA", "8º ANO ALFA"]
ESCOLAS_AMOSTRA = {"CACILDA": 2, "CAIC": 1}  # CACILDA tem 2 formatos => 2 amostras


def _close_overlays(page):
    try:
        page.keyboard.press("Escape")
    except Exception:  # noqa: BLE001, S110
        pass
    time.sleep(0.3)


def _sel(page, label, val):
    el = page.locator(f"label:has-text('{label}')").first
    if not el.count():
        return
    c = el.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    sb = c.locator("dx-select-box").first
    if not sb.count():
        return
    hi = sb.locator("input[type='hidden']").first
    if hi.count() and val.lower() in (hi.get_attribute("value") or "").lower():
        return
    btn = sb.locator("[role='button']").first
    clicked = False
    for attempt in range(3):
        try:
            if btn.count() and btn.is_visible():
                btn.click(timeout=5000)
                clicked = True
                break
        except Exception:  # noqa: BLE001
            _close_overlays(page)
            time.sleep(1)
    if not clicked:
        vi = sb.locator("input[role='combobox']").first
        if vi.count():
            try:
                vi.click(timeout=5000)
            except Exception:  # noqa: BLE001
                return
        else:
            return
    time.sleep(1)
    opt = page.get_by_text(val, exact=True).filter(visible=True).first
    if not opt.count():
        opt = page.locator(".dx-overlay-wrapper .dx-item:visible").filter(has_text=val).first
    if opt.count():
        opt.click()
        time.sleep(0.5)
    else:
        _close_overlays(page)


def _dropdown_options(page, label) -> list[str]:
    el = page.locator(f"label:has-text('{label}')").first
    if not el.count():
        return []
    c = el.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    sb = c.locator("dx-select-box").first
    if not sb.count():
        return []
    btn = sb.locator("[role='button']").first
    try:
        if btn.count() and btn.is_visible():
            btn.click(timeout=5000)
        else:
            sb.click(timeout=5000)
    except Exception:  # noqa: BLE001
        _close_overlays(page)
        return []
    time.sleep(1.5)
    opts = page.locator(".dx-overlay-wrapper .dx-item:visible, [role='option']:visible")
    texts = []
    for i in range(opts.count()):
        t = opts.nth(i).inner_text().strip()
        if t and t not in texts:
            texts.append(t)
    _close_overlays(page)
    return texts


def _filtrar(page):
    btn = page.locator("dx-button:has-text('Filtrar')").first
    if btn.count():
        btn.click()
        page.wait_for_load_state("networkidle")
        time.sleep(3)


def read_evaluations_list(page) -> list[str]:
    ev_names = {"conceito", "avaliação bimestral", "atividade complementar"}
    all_hg = page.locator(".dx-datagrid-headers")
    for gi in range(all_hg.count()):
        hg = all_hg.nth(gi)
        pg = hg.locator("xpath=ancestor::dx-data-grid[1]")
        if not pg.count():
            continue
        dr = pg.locator(".dx-datagrid-rowsview .dx-data-row")
        if dr.count() == 0:
            continue
        fc = dr.first.locator("td")
        texts = [fc.nth(i).inner_text().strip() for i in range(fc.count())]
        if any(t.lower() in ev_names for t in texts if t):
            evals = []
            for i in range(dr.count()):
                cells = dr.nth(i).locator("td")
                for ci in range(cells.count()):
                    t = cells.nth(ci).inner_text().strip()
                    if t:
                        evals.append(t)
                        break
            return evals
    return []


def read_student_grid(page) -> list[dict]:
    students: list[dict] = []
    ev_names = {"conceito", "avaliação bimestral", "atividade complementar"}
    all_hg = page.locator(".dx-datagrid-headers")
    for gi in range(all_hg.count()):
        hg = all_hg.nth(gi)
        hc = hg.locator("td[role='columnheader']")
        name_col = grade_col = sit_col = None
        for ci in range(hc.count()):
            a = (hc.nth(ci).get_attribute("aria-label") or "").lower()
            if "aluno" in a:
                name_col = ci
            elif "nota" in a:
                grade_col = ci
            elif "situa" in a:
                sit_col = ci
        if name_col is None or grade_col is None:
            continue
        pg = hg.locator("xpath=ancestor::dx-data-grid[1]")
        if not pg.count():
            continue
        dr = pg.locator(".dx-datagrid-rowsview .dx-data-row")
        if dr.count() == 0:
            continue
        fc = dr.first.locator("td")
        if fc.count() > name_col and fc.nth(name_col).inner_text().strip().lower() in ev_names:
            continue
        for i in range(dr.count()):
            row = dr.nth(i)
            cells = row.locator("td")
            if cells.count() <= name_col:
                continue
            name = cells.nth(name_col).inner_text().strip()
            if not name or name.lower() in ("aluno", "nome"):
                continue
            sit = cells.nth(sit_col).inner_text().strip() if sit_col is not None and cells.count() > sit_col else None
            inputs = cells.nth(grade_col).locator("input")
            raw = inputs.first.input_value().strip() if inputs.count() else cells.nth(grade_col).inner_text().strip()
            gv = None
            if raw:
                try:
                    gv = float(raw.replace(",", "."))
                except ValueError:
                    gv = None
            students.append({"name": name, "grade": gv, "situacao": sit})
        if students:
            break
    return students


def read_date_field(page) -> str | None:
    date_label = page.locator("label:has-text('Data')").first
    if not date_label.count():
        return None
    container = date_label.locator("xpath=ancestor::div[contains(@class, 'dx-field-item')]")
    datebox = container.locator("dx-date-box, .dx-datebox").first
    if not datebox.count():
        return None
    date_input = datebox.locator("input[role='combobox'], input.dx-texteditor-input, input[type='text']").first
    if not date_input.count():
        return None
    return date_input.input_value().strip() or None


def _open_eval_row(page, ev: str) -> bool:
    """Abre a linha da avaliação com dblclick robusto (scroll + fallback JS)."""
    ec = page.locator(f".dx-data-row td:has-text('{ev}')").first
    if not ec.count():
        return False
    try:
        ec.scroll_into_view_if_needed()
        time.sleep(0.3)
        ec.dblclick(timeout=8000)
        return True
    except Exception:  # noqa: BLE001, S110
        pass
    try:
        ec.dblclick(force=True, timeout=8000)
        return True
    except Exception:  # noqa: BLE001, S110
        pass
    try:
        ec.dispatch_event("dblclick")
        return True
    except Exception:  # noqa: BLE001
        return False


def _norm(name: str) -> str:
    import re
    import unicodedata
    t = "".join(c for c in unicodedata.normalize("NFKD", name) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", t)).strip().lower()


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def read_excel_eh() -> dict:
    """{turma_grp: {norm_name: E+H}} só para as turmas do resto BIMESTRAL."""
    sheets = {"7º ANO B": "7º ANO BETA", "7º ANO G": "7º ANO GAMA", "8º ANO A": "8º ANO ALFA"}
    wb = load_workbook(EXCEL_PATH, data_only=True)
    out = {}
    for sheet, turma in sheets.items():
        ws = wb[sheet]
        name_col = None
        for c in range(1, 5):
            v = ws.cell(3, c).value
            if v and "nome" in str(v).lower():
                name_col = c
                break
        if name_col is None:
            name_col = 1 if sheet == "8º ANO A" else 2
        data = {}
        for r in range(4, (ws.max_row or 3) + 1):
            raw = ws.cell(r, name_col).value
            if raw is None or not str(raw).strip():
                continue
            e = _num(ws.cell(r, 5).value)
            h = _num(ws.cell(r, 8).value)
            if e is not None and h is not None:
                data[_norm(str(raw).strip())] = round(e + h, 2)
        out[turma] = data
    return out


def nav(page, escola, turma, periodo=TARGET_PERIODO):
    page.goto(GRP_URL, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle")
    time.sleep(2)
    _sel(page, "Escola", escola)
    time.sleep(1)
    _sel(page, "Turma", turma)
    time.sleep(1)
    _sel(page, "Disciplina", TARGET_DISCIPLINA)
    time.sleep(1)
    _sel(page, "Período de Avaliação", periodo)
    time.sleep(1)
    _filtrar(page)


def main():
    relatorio: dict = {"coraci_bimestral": {}, "outras_escolas": {}, "aulas": None}
    ts = datetime.now().astimezone().strftime("%Y%m%d-%H%M%S")
    out = Path("artifacts/audit") / f"auditoria-objetiva-{ts}.json"

    def _save():
        out.write_text(json.dumps(relatorio, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[salvo] {out}", flush=True)

    session = BrowserSession.start(headless=False, state_dir="state/playwright")
    try:
        page = session.open_grp()
        navigate_to_grades_page(page)
        time.sleep(2)
        if not page.locator("label:has-text('Escola')").count():
            print("🔑 Login necessário — faça login na janela (até 3 min)...", flush=True)
            for _ in range(36):
                time.sleep(5)
                if page.locator("label:has-text('Escola')").count():
                    break
                try:
                    if "resultadoavaliacao" not in page.url:
                        navigate_to_grades_page(page)
                except Exception:  # noqa: BLE001
                    navigate_to_grades_page(page)
            if not page.locator("label:has-text('Escola')").count():
                print("❌ Login não concluído.")
                return

        # ---- 1. CORACI: BIMESTRAL que falta (7º B, 7º G, 8º A) ----
        excel_eh = read_excel_eh()
        for turma in BIMESTRAL_RESTO:
            try:
                print(f"\n[CORACI] {turma} — AVALIAÇÃO BIMESTRAL", flush=True)
                nav(page, ESCOLA_CORACI, turma)
                if not _open_eval_row(page, "AVALIAÇÃO BIMESTRAL"):
                    relatorio["coraci_bimestral"][turma] = {"erro": "não encontrada/abertura falhou"}
                    print("  não encontrada", flush=True)
                    continue
                time.sleep(3)
                page.wait_for_load_state("networkidle")
                data = read_date_field(page)
                alunos = read_student_grid(page)
                esperado = excel_eh.get(turma, {})
                divergentes, vazios, sem_fonte = [], [], []
                for a in alunos:
                    sit = (a.get("situacao") or "").lower()
                    exp = esperado.get(_norm(a["name"]))
                    if sit.startswith("encerrad"):
                        continue
                    if a["grade"] is None:
                        vazios.append({"aluno": a["name"], "esperado": exp})
                    elif exp is None:
                        sem_fonte.append({"aluno": a["name"], "grp": a["grade"]})
                    elif abs(a["grade"] - exp) > TOL_BIM:
                        divergentes.append({"aluno": a["name"], "grp": a["grade"], "planilha": exp})
                relatorio["coraci_bimestral"][turma] = {
                    "data": data,
                    "divergentes": divergentes,
                    "vazios": vazios,
                    "sem_fonte": sem_fonte,
                    "total_grp": len(alunos),
                }
                print(f"  data={data} | divergentes={len(divergentes)} vazios={len(vazios)} sem_fonte={len(sem_fonte)}", flush=True)
            except Exception as e:  # noqa: BLE001
                relatorio["coraci_bimestral"][turma] = {"erro": str(e)}
                print(f"  ❌ erro: {e}", flush=True)
        _save()

        # ---- 2. Outras escolas: 1 turma-amostra por escola ----
        escolas = _dropdown_options(page, "Escola") if False else None  # evita dropdown aberto
        page.goto(GRP_URL, wait_until="domcontentloaded")
        page.wait_for_load_state("networkidle")
        time.sleep(2)
        escolas = _dropdown_options(page, "Escola")
        print(f"\nESCOLAS: {escolas}", flush=True)
        for escola in escolas:
            ekey = escola.upper()
            if "CORACI" in ekey:
                continue
            n_amostras = None
            for k, n in ESCOLAS_AMOSTRA.items():
                if k in ekey:
                    n_amostras = n
                    break
            if n_amostras is None:
                print(f"\n[skip] {escola}", flush=True)
                continue
            try:
                _sel(page, "Escola", escola)
                time.sleep(2)
                turmas = _dropdown_options(page, "Turma")
                print(f"\n##### {escola} — turmas: {turmas}", flush=True)
                if not turmas:
                    relatorio["outras_escolas"][escola] = {"erro": "sem turmas listadas"}
                    continue
                # amostras espalhadas: primeira, última (e prefere 8º/9º quando 1 amostra)
                amostras = []
                if n_amostras == 1:
                    amostra = turmas[0]
                    for t in turmas:
                        if "9º" in t or "8º" in t:
                            amostra = t
                            break
                    amostras = [amostra]
                else:
                    amostras = [turmas[0]]
                    for t in reversed(turmas):
                        if t not in amostras:
                            amostras.append(t)
                            break
                info = {"todas_turmas": turmas, "amostras": {}}
                for amostra in amostras:
                    print(f"\n  >>> amostra: {amostra}", flush=True)
                    nav(page, escola, amostra)
                    disp = read_evaluations_list(page)
                    a_info = {"avaliacoes_disponiveis": disp, "avaliacoes": {}}
                    print(f"  avaliações: {disp}", flush=True)
                    for ev in disp:
                        if not _open_eval_row(page, ev):
                            a_info["avaliacoes"][ev] = {"erro": "falha ao abrir"}
                            continue
                        time.sleep(3)
                        page.wait_for_load_state("networkidle")
                        d = read_date_field(page)
                        alunos = read_student_grid(page)
                        com = sum(1 for a in alunos if a["grade"] is not None)
                        a_info["avaliacoes"][ev] = {
                            "data": d,
                            "total": len(alunos),
                            "com_nota": com,
                            "sem_nota": len(alunos) - com,
                        }
                        print(f"    {ev}: data={d} total={len(alunos)} com_nota={com}", flush=True)
                        nav(page, escola, amostra)  # volta para a lista
                    info["amostras"][amostra] = a_info
                relatorio["outras_escolas"][escola] = info
            except Exception as e:  # noqa: BLE001
                relatorio["outras_escolas"][escola] = {"erro": str(e)}
                print(f"  ❌ erro na escola: {e}", flush=True)
            _save()

        # ---- 3. Mudar de campo: explorar AULAS ----
        # ---- 3. Mudar de campo: explorar AULAS ----
        try:
            print("\n##### EXPLORANDO ÁREA DE AULAS #####", flush=True)
            page.goto(GRP_HOME, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")
            time.sleep(2)
            links = page.locator("a")
            achados = []
            for i in range(links.count()):
                try:
                    txt = links.nth(i).inner_text().strip()
                    href = links.nth(i).get_attribute("href") or ""
                except Exception:  # noqa: BLE001, S112
                    continue
                if txt and ("aula" in txt.lower() or "aula" in href.lower()):
                    achados.append({"texto": txt, "href": href})
            relatorio["aulas"] = {"links_encontrados": achados}
            print(f"  links com 'aula': {achados}", flush=True)
            # tenta abrir o primeiro link de aula
            if achados:
                alvo = page.get_by_text(achados[0]["texto"], exact=True).first
                if alvo.count():
                    try:
                        alvo.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(3)
                        relatorio["aulas"]["url_aberta"] = page.url
                        corpo = page.locator("body").inner_text()[:3000]
                        relatorio["aulas"]["texto_pagina"] = corpo
                        print(f"  página aberta: {page.url}", flush=True)
                        print("  conteúdo (início):", flush=True)
                        print(corpo[:1200], flush=True)
                    except Exception as e:  # noqa: BLE001
                        relatorio["aulas"]["erro"] = str(e)
        except Exception as e:  # noqa: BLE001
            relatorio["aulas"] = {"erro": str(e)}
            print(f"  ❌ erro na fase AULAS: {e}", flush=True)
        _save()
    finally:
        session.close()
        _save()
    print("\nFIM.", flush=True)


if __name__ == "__main__":
    main()
