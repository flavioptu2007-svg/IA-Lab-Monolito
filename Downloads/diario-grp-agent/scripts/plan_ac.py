#!/usr/bin/env python3
"""Plano OFFLINE de lançamento ATIVIDADE COMPLEMENTAR (coluna G).

Cruza o mapa read-only do GRP (artifacts/audit/ac-mapping-*.json) com a
coluna G da planilha e classifica cada aluno:
  LANCAR      -> casado no GRP, campo vazio, G numérico e <= escala (10)
  BLOQUEADO   -> erro de fórmula / sem valor / acima da escala / já com nota /
                 não encontrado / ambíguo / divergente

Não abre navegador, não toca no GRP.
"""
from __future__ import annotations

import glob
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from grp_agent.student_matcher import MatchClassification, StudentMatcher
from openpyxl import load_workbook

EXCEL_PATH = "/home/flavio/Trabalho/Secretária/PROVAS 2026 2/Notas AGT e Ciclo 2 - Trimestre 2.xlsx"
ARREDONDAR_CASAS = 2  # decisão do usuário: 2 casas (ex.: 7.969 -> 7,97; GRP já mostra 7.97 p/ ADRYAN)

TURMA_MAP = {
    "6º ANO A": "6º ANO ALFA",
    "6º ANO B": "6º ANO BETA",
    "6º ANO G": "6º ANO GAMA",
    "7º ANO A": "7º ANO ALFA",
    "7º ANO B": "7º ANO BETA",
    "7º ANO G": "7º ANO GAMA",
    "8º ANO A": "8º ANO ALFA",
}


def read_column(sheet_name: str, col_idx: int) -> list[dict]:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[sheet_name]
    name_col = None
    for c in range(1, 5):
        v = ws.cell(3, c).value
        if v and "nome" in str(v).lower():
            name_col = c
            break
    if name_col is None:
        name_col = 1 if sheet_name == "8º ANO A" else 2
    rows = []
    for r in range(4, (ws.max_row or 3) + 1):
        raw = ws.cell(r, name_col).value
        if raw is None or not str(raw).strip():
            continue
        v = ws.cell(r, col_idx).value
        rows.append({"name": str(raw).strip(), "g_raw": v})
    return rows


def classify_value(g_raw):
    """Retorna (valor_float|None, motivo_bloqueio|None)."""
    if g_raw is None:
        return None, "sem_valor"
    if isinstance(g_raw, str):
        s = g_raw.strip()
        if s.startswith("#"):
            return None, "erro_formula"
        try:
            return float(s.replace(",", ".")), None
        except ValueError:
            return None, "nao_numerico"
    if isinstance(g_raw, (int, float)):
        return float(g_raw), None
    return None, "nao_numerico"


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Plano offline de lançamento")
    parser.add_argument("--coluna", default="G", help="Coluna da nota (letra)")
    parser.add_argument("--escala", type=float, default=10.0, help="Escala máxima da nota")
    parser.add_argument("--avaliacao", default="ATIVIDADE COMPLEMENTAR")
    parser.add_argument("--mapping", default=None, help="Glob do mapa GRP (default: deriva da avaliação)")
    parser.add_argument("--out", default="artifacts/audit/ac-plan.json")
    parser.add_argument("--excluir", default="", help="Turmas GRP a excluir, separadas por vírgula")
    args = parser.parse_args()

    col_idx = ord(args.coluna.upper()) - ord("A") + 1
    escala_max = args.escala
    excluidas = {t.strip() for t in args.excluir.split(",") if t.strip()}
    slug = args.avaliacao.replace(" ", "_").lower()
    mapping_glob = args.mapping or f"artifacts/audit/{slug}-mapping-*.json"
    mapping_files = sorted(glob.glob(mapping_glob))
    if not mapping_files:
        print(f"Nenhum mapa em {mapping_glob}; rode scripts/readonly_map_ac.py --avaliacao '{args.avaliacao}'.")
        return
    mapping = json.loads(Path(mapping_files[-1]).read_text(encoding="utf-8"))
    print(f"Avaliação: {args.avaliacao} | coluna {args.coluna.upper()} | escala {escala_max:g}")
    print(f"Mapa GRP: {mapping_files[-1]}\n")

    matcher = StudentMatcher()
    plan = {}
    tot = {"lancar": 0, "bloqueado": 0}
    for sheet, turma in TURMA_MAP.items():
        if turma in excluidas:
            continue
        grp = mapping.get(turma, {})
        grp_students = grp.get("alunos", [])
        excel_rows = read_column(sheet, col_idx)
        items = []
        for row in excel_rows:
            valor, motivo = classify_value(row["g_raw"])
            item = {
                "nome_planilha": row["name"],
                "g_raw": row["g_raw"] if isinstance(row["g_raw"], (int, float, type(None))) else str(row["g_raw"]),
                "valor": valor,
                "motivo_bloqueio": motivo,
                "status": "BLOQUEADO" if motivo else None,
            }
            if motivo is None and valor is not None and valor > escala_max:
                item["motivo_bloqueio"] = "acima_da_escala"
                item["status"] = "BLOQUEADO"

            mr = matcher.match(row["name"], grp_students)
            item["match"] = mr.grp_name
            item["classificacao"] = mr.classification.value
            item["confianca"] = round(mr.confidence, 2)

            if item["status"] is None:
                grp_entry = next(
                    (s for s in grp_students if s["name"] == mr.grp_name), None
                )
                sit = (grp_entry or {}).get("situacao")
                if sit and sit.strip().lower() in ("encerrada", "encerrado"):
                    item["status"], item["motivo_bloqueio"] = "BLOQUEADO", "matricula_encerrada"
                elif mr.classification in (MatchClassification.NAO_ENCONTRADA,):
                    item["status"], item["motivo_bloqueio"] = "BLOQUEADO", "nao_encontrado"
                elif mr.classification == MatchClassification.AMBIGUA:
                    item["status"], item["motivo_bloqueio"] = "BLOQUEADO", "ambiguo"
                else:
                    nota_grp = grp_entry["grade"] if grp_entry else None
                    item["nota_grp_atual"] = nota_grp
                    lancada = round(float(f"{valor:.{ARREDONDAR_CASAS}f}"), ARREDONDAR_CASAS)
                    item["valor_lancar"] = lancada
                    if nota_grp is not None:
                        if abs(nota_grp - lancada) <= 0.05:
                            item["status"], item["motivo_bloqueio"] = "OK_JA_LANCADA", None
                        else:
                            item["status"], item["motivo_bloqueio"] = "BLOQUEADO", "ja_com_nota_divergente"
                    else:
                        item["status"] = "LANCAR"
            items.append(item)

        n_lancar = sum(1 for i in items if i["status"] == "LANCAR")
        n_ok = sum(1 for i in items if i["status"] == "OK_JA_LANCADA")
        n_bloq = sum(1 for i in items if i["status"] == "BLOQUEADO")
        tot["lancar"] += n_lancar
        tot["bloqueado"] += n_bloq
        plan[turma] = {
            "sheet": sheet,
            "data_campo_grp": grp.get("data_campo"),
            "total_planilha": len(items),
            "a_lancar": n_lancar,
            "ja_lancadas_ok": n_ok,
            "bloqueados": n_bloq,
            "itens": items,
        }
        print(f"=== {turma} (aba {sheet}) — data no GRP: {grp.get('data_campo') or 'SEM DATA'}")
        print(f"    planilha: {len(items)} | A LANÇAR: {n_lancar} | já OK: {n_ok} | bloqueados: {n_bloq}")
        for i in items:
            if i["status"] == "LANCAR":
                print(f"    ✅ {i['nome_planilha']:42s} G={i['valor']:.3f} -> lançar {i['valor_lancar']} ({i['classificacao']})")
            elif i["status"] == "OK_JA_LANCADA":
                print(f"    ✔  {i['nome_planilha']:42s} já consta {i['nota_grp_atual']} no GRP (= planilha)")
            else:
                print(f"    ⛔ {i['nome_planilha']:42s} G={i['g_raw']!r} -> {i['motivo_bloqueio']}"
                      + (f" (GRP={i.get('nota_grp_atual')})" if i.get('nota_grp_atual') is not None else ""))
        print()

    print(f"TOTAL: {tot['lancar']} a lançar | {tot['bloqueado']} bloqueados")
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Plano salvo em: {out}")


if __name__ == "__main__":
    main()
